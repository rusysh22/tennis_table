import os
import secrets
import threading
import time
import hashlib
import json
from collections import Counter, deque
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import unquote, urlsplit
from zoneinfo import ZoneInfo

from flask import (
    Flask, render_template, request, redirect, url_for, session, jsonify, abort, flash, send_file
)
import io
import uuid
from PIL import Image, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash
from dotenv import load_dotenv

import og_image
import utils
import license_client

load_dotenv()

app = Flask(__name__)


def _env_flag(name, default=False):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}

SECRET_PATH = os.path.join(os.path.dirname(__file__), ".secret_key")
app.secret_key = os.environ.get("SECRET_KEY")
if not app.secret_key:
    try:
        if os.path.exists(SECRET_PATH):
            with open(SECRET_PATH, "r") as f:
                app.secret_key = f.read().strip()
    except OSError:
        pass  # filesystem tidak bisa dibaca (mis. serverless) - lanjut ke fallback di bawah
if not app.secret_key:
    app.secret_key = os.urandom(24).hex()
    try:
        with open(SECRET_PATH, "w") as f:
            f.write(app.secret_key)
    except OSError:
        pass  # filesystem read-only (mis. Vercel) - key tetap dipakai, cuma tidak tersimpan
               # ke disk sehingga bisa beda tiap cold start (sesi admin akan ter-logout).
               # Set env var SECRET_KEY di hosting supaya key konsisten antar-instance.

app.config.update(
    ADMIN_PASSWORD=os.environ.get("ADMIN_PASSWORD"),
    ADMIN_PASSWORD_HASH=os.environ.get("ADMIN_PASSWORD_HASH"),
    ADMIN_PASSWORD_GALLERY=os.environ.get("ADMIN_PASSWORD_GALLERY"),
    ADMIN_PASSWORD_HASH_GALLERY=os.environ.get("ADMIN_PASSWORD_HASH_GALLERY"),
    ADMIN_PASSWORD_TABLE_TENNIS=os.environ.get("ADMIN_PASSWORD_TABLE_TENNIS"),
    ADMIN_PASSWORD_HASH_TABLE_TENNIS=os.environ.get("ADMIN_PASSWORD_HASH_TABLE_TENNIS"),
    ADMIN_PASSWORD_PADEL=os.environ.get("ADMIN_PASSWORD_PADEL"),
    ADMIN_PASSWORD_HASH_PADEL=os.environ.get("ADMIN_PASSWORD_HASH_PADEL"),
    ADMIN_PASSWORD_BADMINTON=os.environ.get("ADMIN_PASSWORD_BADMINTON"),
    ADMIN_PASSWORD_HASH_BADMINTON=os.environ.get("ADMIN_PASSWORD_HASH_BADMINTON"),
    MAX_CONTENT_LENGTH=int(os.environ.get("MAX_UPLOAD_BYTES", 48 * 1024 * 1024)),
    PERMANENT_SESSION_LIFETIME=timedelta(
        hours=int(os.environ.get("ADMIN_SESSION_HOURS", "8"))
    ),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=_env_flag("SESSION_COOKIE_SECURE", default=True),
)

# Admin accounts: each operator (general / gallery / one per sport) gets its own
# password via environment variables, following the same ADMIN_PASSWORD_HASH
# convention as before. An account only shows up in the login dropdown if its
# env var is actually set -- a deployment that only sets ADMIN_PASSWORD_HASH
# (today's setup) keeps working unchanged, just with a single-option dropdown.
ADMIN_ACCOUNTS = [
    {"key": "general", "label": "General (Semua Cabor)", "role": "general", "sport_key": None,
     "password_env": "ADMIN_PASSWORD", "hash_env": "ADMIN_PASSWORD_HASH"},
    {"key": "gallery", "label": "Admin Galeri", "role": "gallery", "sport_key": None,
     "password_env": "ADMIN_PASSWORD_GALLERY", "hash_env": "ADMIN_PASSWORD_HASH_GALLERY"},
    {"key": "table-tennis", "label": "Admin Tenis Meja", "role": "sport", "sport_key": "table-tennis",
     "password_env": "ADMIN_PASSWORD_TABLE_TENNIS", "hash_env": "ADMIN_PASSWORD_HASH_TABLE_TENNIS"},
    {"key": "padel", "label": "Admin Padel", "role": "sport", "sport_key": "padel",
     "password_env": "ADMIN_PASSWORD_PADEL", "hash_env": "ADMIN_PASSWORD_HASH_PADEL"},
    {"key": "badminton", "label": "Admin Badminton", "role": "sport", "sport_key": "badminton",
     "password_env": "ADMIN_PASSWORD_BADMINTON", "hash_env": "ADMIN_PASSWORD_HASH_BADMINTON"},
]

Image.MAX_IMAGE_PIXELS = int(os.environ.get("MAX_IMAGE_PIXELS", "25000000"))

# Media (champion photos, gallery, match documents) lives on local disk under
# static/uploads, which is a persistent Docker volume -- no external object
# storage involved.
LOCAL_UPLOAD_ROOT = os.path.join(app.root_path, "static", "uploads")


def compress_and_upload_image(file, match_id):
    """Compress an uploaded image to JPEG and save it under static/uploads."""
    try:
        img = Image.open(file)
        img.verify()
        file.seek(0)
        img = Image.open(file)
        img = ImageOps.exif_transpose(img)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        # Compress the image
        img.thumbnail((1200, 1200))  # Resize if larger than 1200px
        output = io.BytesIO()
        img.save(output, format="JPEG", quality=75, optimize=True)
        output.seek(0)

        filename = f"docs/{match_id}_{uuid.uuid4().hex[:8]}.jpg"
        dest_path = os.path.join(LOCAL_UPLOAD_ROOT, filename)
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(dest_path, "wb") as f:
            f.write(output.getbuffer())

        file_url = url_for("static", filename=f"uploads/{filename}", _external=False)
        return file_url, None
    except Exception as e:
        return None, str(e)


def _local_upload_path(file_url):
    """Resolve a static/uploads URL (as returned by compress_and_upload_image)
    back to its on-disk path, or None if it doesn't point into that tree."""
    if not file_url:
        return None
    path = urlsplit(file_url).path
    prefix = url_for("static", filename="uploads/", _external=False)
    if not path.startswith(prefix):
        return None
    relative = unquote(path[len(prefix):])
    full_path = os.path.normpath(os.path.join(LOCAL_UPLOAD_ROOT, relative))
    if not full_path.startswith(os.path.normpath(LOCAL_UPLOAD_ROOT) + os.sep):
        return None
    return full_path


def delete_from_supabase(file_url):
    """Delete a previously uploaded file from local disk (name kept for
    call-site compatibility with the old S3-backed implementation)."""
    path = _local_upload_path(file_url)
    if not path:
        return
    try:
        os.remove(path)
    except OSError:
        pass


GALLERY_UPLOAD_BATCH_LIMIT = 10

ROUND_LABELS = {1: "Babak 1", 2: "Babak 2", 3: "Babak 3", 4: "Final"}
STATUS_LABELS = {
    "draft": "Draf",
    "scheduled": "Terjadwal",
    "check_in": "Check-in",
    "live": "Live",
    "completed": "Selesai",
    "postponed": "Ditunda",
    "suspended": "Ditangguhkan",
    "cancelled": "Dibatalkan",
}


# ---------- helpers ----------

TBD_COLOR = "#9c9c9c"
TBD_TEXT = "#ffffff"
CSRF_SESSION_KEY = "_csrf_token"
_LOGIN_ATTEMPTS = {}
_LOGIN_LOCKED_UNTIL = {}
_LOGIN_RATE_LOCK = threading.Lock()

app.config.setdefault("LOGIN_MAX_ATTEMPTS", 5)
app.config.setdefault("LOGIN_WINDOW_SECONDS", 15 * 60)
app.config.setdefault("LOGIN_LOCK_SECONDS", 15 * 60)

_RATE_BUCKETS = {}
_RATE_LOCK = threading.Lock()


def _rate_limited(key, max_events, window_seconds):
    """True if `key` has hit max_events within window_seconds; records this
    call as an event either way, so only call this once per attempt, right
    before acting on it."""
    now = time.monotonic()
    with _RATE_LOCK:
        bucket = _RATE_BUCKETS.setdefault(key, deque())
        while bucket and bucket[0] <= now - window_seconds:
            bucket.popleft()
        if len(bucket) >= max_events:
            return True
        bucket.append(now)
        return False


def get_csrf_token():
    token = session.get(CSRF_SESSION_KEY)
    if not token:
        token = secrets.token_urlsafe(32)
        session[CSRF_SESSION_KEY] = token
    return token


def _safe_internal_next(value):
    if not value:
        return None
    parsed = urlsplit(value)
    if parsed.scheme or parsed.netloc or not value.startswith("/"):
        return None
    decoded = unquote(value)
    if (
        value.startswith("//")
        or "\\" in decoded
        or "\r" in decoded
        or "\n" in decoded
    ):
        return None
    return value


def _configured_accounts():
    return [
        account for account in ADMIN_ACCOUNTS
        if app.config.get(account["hash_env"]) or app.config.get(account["password_env"])
    ]


def _verify_account_password(account, candidate):
    password_hash = app.config.get(account["hash_env"])
    if password_hash:
        try:
            return check_password_hash(password_hash, candidate)
        except ValueError:
            return False
    password = app.config.get(account["password_env"])
    return bool(password) and secrets.compare_digest(password, candidate)


def _assert_sport_access(sport_key):
    """Abort 403 if the logged-in sport-scoped admin doesn't own this sport_key."""
    if session.get("admin_role") == "sport" and session.get("admin_sport") != (sport_key or "table-tennis"):
        abort(403)


def _login_client_key():
    return request.remote_addr or "unknown"


def _login_lock_remaining(client_key):
    now = time.monotonic()
    with _LOGIN_RATE_LOCK:
        locked_until = _LOGIN_LOCKED_UNTIL.get(client_key, 0)
        if locked_until <= now:
            _LOGIN_LOCKED_UNTIL.pop(client_key, None)
            return 0
        return int(locked_until - now) + 1


def _record_login_failure(client_key):
    now = time.monotonic()
    window = app.config["LOGIN_WINDOW_SECONDS"]
    with _LOGIN_RATE_LOCK:
        attempts = _LOGIN_ATTEMPTS.setdefault(client_key, deque())
        while attempts and attempts[0] <= now - window:
            attempts.popleft()
        attempts.append(now)
        if len(attempts) >= app.config["LOGIN_MAX_ATTEMPTS"]:
            _LOGIN_LOCKED_UNTIL[client_key] = now + app.config["LOGIN_LOCK_SECONDS"]
            attempts.clear()


def _clear_login_failures(client_key):
    with _LOGIN_RATE_LOCK:
        _LOGIN_ATTEMPTS.pop(client_key, None)
        _LOGIN_LOCKED_UNTIL.pop(client_key, None)


@app.before_request
def protect_against_csrf():
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    expected = session.get(CSRF_SESSION_KEY)
    supplied = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not expected or not supplied or not secrets.compare_digest(expected, supplied):
        abort(400, description="Token CSRF tidak valid atau sudah kedaluwarsa.")
    return None


@app.before_request
def enforce_license_interception():
    if app.config.get("TESTING") and not app.config.get("TEST_LICENSE_INTERCEPT"):
        return None

    allowed_endpoints = {
        "static", "admin_login", "admin_logout",
        "admin_license", "admin_activate_license",
        "admin_validate_license", "admin_deactivate_license",
        "license_lockout", "not_found"
    }
    endpoint = request.endpoint or ""
    if endpoint in allowed_endpoints or endpoint.startswith("static"):
        return None

    is_active, reason = license_client.is_license_active()
    if not is_active:
        if session.get("is_admin") and session.get("admin_role", "general") == "general":
            flash(f"⚠️ Akses ditahan: {reason} Silakan aktifkan lisensi dari Berlanggan.web.id.", "error")
            return redirect(url_for("admin_license"))
        return redirect(url_for("license_lockout", reason=reason))
    return None


GENERAL_ONLY_ENDPOINTS = {
    "admin_sites", "admin_save_site", "admin_delete_site",
    "admin_participants", "admin_save_participant", "admin_delete_participant", "admin_shuffle_groups",
    "admin_categories", "admin_save_category",
    "admin_rules", "admin_save_rules",
    "admin_schedule_generator", "admin_generate_group_to_knockout", "admin_sync_knockout",
    "admin_settings", "admin_announcement", "admin_live_stream",
    "admin_utilities", "admin_reset", "admin_shuffle_putra", "admin_shuffle_campuran",
    "admin_license", "admin_activate_license", "admin_validate_license", "admin_deactivate_license",
    "admin_delete_live_chat_message",
}
GALLERY_ENDPOINTS = {"admin_upload_gallery_photos", "admin_delete_gallery_photo"}


@app.before_request
def enforce_admin_role_scope():
    if not request.path.startswith("/admin") or not session.get("is_admin"):
        return None
    endpoint = request.endpoint or ""
    if endpoint in {"admin_login", "admin_logout"}:
        return None
    role = session.get("admin_role", "general")
    if role == "gallery" and endpoint not in GALLERY_ENDPOINTS:
        flash("Akun ini hanya memiliki akses ke menu Galeri.", "error")
        return redirect(url_for("galeri"))
    if role == "sport" and (endpoint in GENERAL_ONLY_ENDPOINTS or endpoint in GALLERY_ENDPOINTS):
        flash("Menu ini tidak tersedia untuk akun cabor Anda.", "error")
        return redirect(url_for("admin_dashboard"))
    return None


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(), microphone=(), geolocation=(), payment=(), usb=()",
    )
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; base-uri 'self'; object-src 'none'; frame-ancestors 'none'; "
        "frame-src 'self' https://www.youtube.com https://www.youtube-nocookie.com https://player.vimeo.com; "
        "form-action 'self'; img-src 'self' data: https://*.ytimg.com https://*.youtube.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "script-src 'self' 'unsafe-inline' https://www.googletagmanager.com; "
        "connect-src 'self' https://www.google-analytics.com https://*.google-analytics.com",
    )
    if request.is_secure:
        response.headers.setdefault(
            "Strict-Transport-Security", "max-age=31536000; includeSubDomains"
        )
    if request.path.startswith("/admin"):
        response.headers.setdefault("Cache-Control", "no-store")
    return response


def champion_display(teams, code):
    """Data tampilan juara: chip kode tim + nama anggota terpisah (bukan digabung '/')."""
    if not code:
        return None
    t = teams.get(code, {})
    return {
        "code": t.get("code", code),
        "player1": t.get("player1", ""),
        "player2": t.get("player2", ""),
        "color": t.get("color", TBD_COLOR),
        "text": t.get("text", TBD_TEXT),
    }


def enrich_match(m, teams):
    m = dict(m)
    m.setdefault("sport_key", "table-tennis")
    m.setdefault(
        "stage_type", "final" if m.get("group") == "FINAL" else ("semifinal" if m.get("group") in ("KNOCKOUT", "SEMIFINAL") else "group")
    )
    m.setdefault(
        "stage_key", "final" if m.get("group") == "FINAL" else ("semifinal" if m.get("group") in ("KNOCKOUT", "SEMIFINAL") else "group-stage")
    )
    m.setdefault("stage_label", m.get("round_label", "Pertandingan"))
    m["team_a_label"] = utils.team_label(teams, m["team_a"]) if m.get("team_a") else (m.get("qualification_slot_a") or "Menunggu Kontestan")
    m["team_b_label"] = utils.team_label(teams, m["team_b"]) if m.get("team_b") else (m.get("qualification_slot_b") or "Menunggu Kontestan")
    ta = teams.get(m["team_a"], {}) if m.get("team_a") else {}
    tb = teams.get(m["team_b"], {}) if m.get("team_b") else {}
    m["team_a_code"] = utils.team_short(teams, m["team_a"]) if m.get("team_a") else (m.get("qualification_slot_a") or "TBD")
    m["team_b_code"] = utils.team_short(teams, m["team_b"]) if m.get("team_b") else (m.get("qualification_slot_b") or "TBD")
    m["team_a_color"] = ta.get("color", TBD_COLOR)
    m["team_a_text"] = ta.get("text", TBD_TEXT)
    m["team_b_color"] = tb.get("color", TBD_COLOR)
    m["team_b_text"] = tb.get("text", TBD_TEXT)
    m["team_a_player1"] = ta.get("player1", m.get("qualification_slot_a") or "TBD")
    m["team_a_player2"] = ta.get("player2", "")
    m["team_b_player1"] = tb.get("player1", m.get("qualification_slot_b") or "TBD")
    m["team_b_player2"] = tb.get("player2", "")
    m["date_label"] = (
        utils.format_date_id(m["date"]) if m.get("date") else "Belum dijadwalkan"
    )
    m["status_label"] = STATUS_LABELS.get(m["status"], m["status"])
    m["is_walkover"] = bool(m.get("walkover"))
    is_capsa = m.get("sport_key") == "capsa-susun"
    if is_capsa:
        sa = sum(s[0] for s in m["sets"]) if m["sets"] else 0
        sb = sum(s[1] for s in m["sets"]) if m["sets"] else 0
    else:
        sa, sb = utils.compute_sets_won(m["sets"]) if m["sets"] else (0, 0)
    m["sets_a"] = sa
    m["sets_b"] = sb
    if is_capsa:
        m["sets_needed"] = 5
        m["best_of_label"] = "5 Ronde — Total Poin Terendah Menang"
        m["segment_term"] = "ronde"
        m["segment_label"] = "Ronde"
        m["scoring_unit_label"] = "Poin"
    else:
        m["sets_needed"] = utils.sets_needed_to_win(m)
        m["best_of_label"] = f"Best of {m['sets_needed'] * 2 - 1}"
        score_terms = utils.scorekeeper_terms(m)
        m["segment_term"] = score_terms["segment_term"]
        m["segment_label"] = score_terms["segment_label"]
        m["scoring_unit_label"] = score_terms["unit_label"]
    _, score_errors = utils.validate_recorded_result(m)
    m["score_errors"] = score_errors
    m["score_is_valid"] = not score_errors
    
    stored_winner = m.get("winner")
    m["winner"] = None
    if m["status"] == "completed":
        if stored_winner == m.get("team_a") or sa > sb:
            m["winner"] = "a"
        elif stored_winner == m.get("team_b") or sb > sa:
            m["winner"] = "b"
        
    m["comments"] = m.get("comments", [])
    
    # Voting logic
    match_dt_str = f"{m.get('date', '')} {m.get('time', '')}"
    try:
        match_dt = datetime.strptime(match_dt_str, "%Y-%m-%d %H:%M").replace(
            tzinfo=ZoneInfo(
                m.get("_timezone")
                or os.environ.get("TOURNAMENT_TIMEZONE", "Asia/Jakarta")
            )
        )
        m["is_voting_open"] = utils.now_wib(
            m.get("_timezone")
        ) < match_dt - timedelta(hours=1)
    except:
        m["is_voting_open"] = False
        
    if m["status"] == "completed":
        m["is_voting_open"] = False
        
    m["votes"] = m.get("votes", {"a": {}, "b": {}})
    total_a = sum(m["votes"]["a"].values())
    total_b = sum(m["votes"]["b"].values())
    total_votes = total_a + total_b
    
    m["vote_total_a"] = total_a
    m["vote_total_b"] = total_b
    if total_votes > 0:
        m["vote_pct_a"] = int(round((total_a / total_votes) * 100))
        m["vote_pct_b"] = 100 - m["vote_pct_a"]
    else:
        m["vote_pct_a"] = 0
        m["vote_pct_b"] = 0
        
    return m


def competition_context(matches, teams, config, sport_key=None):
    competition = utils.build_competition_view(
        matches, teams, config, sport_key=sport_key
    )
    for division in competition.get("divisions", []):
        division["champion_team"] = champion_display(
            teams, division.get("champion")
        )
        division["silver_team"] = champion_display(teams, division.get("silver"))
        division["bronze_team"] = champion_display(teams, division.get("bronze"))
        division["has_elimination_stage"] = any(
            stage.get("type") not in {"group", "round_robin"}
            for stage in division.get("stages_view", [])
        )
        policy_config = (
            division.get("standing_policy") or {}
        ).get("config") or utils.DEFAULT_STANDING_POLICY
        division["standing_summary"] = (
            f"Menang {policy_config.get('win_points', 2)} poin · "
            f"kalah bertanding {policy_config.get('played_loss_points', 1)} poin · "
            f"kalah W.O. {policy_config.get('walkover_loss_points', 0)} poin"
        )
        for stage in division.get("stages_view", []):
            stage["matches_enriched"] = [
                enrich_match(match, teams) for match in stage.get("matches", [])
            ]
            stage["qualifiers"] = [
                {
                    "group": group_key,
                    "code": code,
                    "team": champion_display(teams, code),
                }
                for group_key, code in sorted(
                    division.get("group_winners", {}).items()
                )
            ]
            for group in stage.get("groups_view", []):
                group["champion_team"] = champion_display(
                    teams, group.get("champion")
                )
                for row in group.get("rows", []):
                    row["display_code"] = utils.team_short(teams, row["code"])
    return competition


def qualification_context(match, matches, teams, config):
    competition = competition_context(matches, teams, config)
    division = next(
        (
            item for item in competition.get("divisions", [])
            if item.get("key") == match.get("category")
            and item.get("sport_key") == match.get("sport_key", "table-tennis")
        ),
        None,
    )
    qualifiers = []
    if division:
        qualifiers = [
            {
                "group": group_key,
                "code": utils.team_short(teams, code),
                "value": code,
            }
            for group_key, code in sorted(division.get("group_winners", {}).items())
        ]
    eligible_teams = {
        code: {**team, "label": utils.team_label(teams, code)}
        for code, team in teams.items()
        if team.get("category") == match.get("category")
        and team.get("sport_key", "table-tennis")
        == match.get("sport_key", "table-tennis")
    }
    return {
        "qualifiers": qualifiers,
        "eligible_teams": eligible_teams,
        "is_elimination": match.get("stage_type") not in {
            None, "group", "round_robin"
        } or match.get("group") == "FINAL",
    }


def scorekeeper_document(match, teams):
    enriched = enrich_match(match, teams)
    state = utils.scorekeeper_state(match)
    event_labels = {
        "score.match_started": "Pertandingan dimulai",
        "score.point_awarded": "Skor ditambahkan",
        "score.undo": "Skor terakhir di-undo",
        "score.match_finished": "Hasil dikonfirmasi selesai",
        "score.correction_opened": "Hasil dibuka untuk koreksi",
    }
    history = []
    for event in reversed(match.get("scorekeeper_events") or []):
        side = event.get("side")
        side_code = (
            enriched.get("team_a_code") if side == "a"
            else enriched.get("team_b_code") if side == "b"
            else None
        )
        history.append(
            {
                "sequence": event.get("sequence"),
                "label": event_labels.get(event.get("event_type"), event.get("event_type")),
                "side_code": side_code,
                "at": event.get("at"),
                "reason": (event.get("metadata") or {}).get("reason"),
            }
        )
        if len(history) >= 8:
            break
    return {
        "match": enriched,
        "state": state,
        "history": history,
        "version": int(match.get("version", 0)),
        "can_start": match.get("status") in {"scheduled", "check_in", "postponed"},
        "can_open_correction": (
            match.get("status") == "completed"
            and not match.get("walkover")
            and bool(match.get("sets"))
        ),
    }


@app.template_filter("truncate_words")
def truncate_words(text, max_words=2):
    return utils.truncate_words(text, max_words)


@app.template_filter("format_datetime_id")
def format_datetime_id(value):
    return utils.format_datetime_id(value)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def _expected_match_version():
    raw_version = request.form.get("version")
    if raw_version is None:
        raise ValueError("Versi pertandingan tidak tersedia. Muat ulang halaman lalu coba lagi.")
    try:
        return int(raw_version)
    except (TypeError, ValueError) as exc:
        raise ValueError("Versi pertandingan tidak valid. Muat ulang halaman.") from exc


def _update_match_or_flash(match_id, expected_version, updater, success_message):
    try:
        updated = utils.update_match(
            match_id, updater, expected_version=expected_version
        )
    except utils.MatchVersionConflictError:
        flash(
            "Pertandingan telah diubah oleh pengguna lain. Muat ulang halaman dan periksa skor terbaru.",
            "error",
        )
        return None
    except (utils.MatchNotFoundError, ValueError) as exc:
        flash(str(exc) or "Pertandingan tidak ditemukan.", "error")
        return None
    flash(success_message, "success")
    if updated and isinstance(updated, dict) and updated.get("category"):
        try:
            utils.auto_seed_knockout(updated["category"])
        except Exception:
            pass
    return updated


def _parse_score_form(max_games, segment_label="Game"):
    games = []
    errors = []
    gap_found = False
    for index in range(1, max_games + 1):
        raw_a = request.form.get(f"set{index}_a", "").strip()
        raw_b = request.form.get(f"set{index}_b", "").strip()
        if not raw_a and not raw_b:
            gap_found = True
            continue
        if not raw_a or not raw_b:
            errors.append(f"Skor {segment_label} {index} harus diisi untuk kedua tim.")
            continue
        if gap_found:
            errors.append(
                f"{segment_label} harus diisi berurutan tanpa melewati "
                f"{segment_label.lower()} kosong."
            )
        try:
            games.append([int(raw_a), int(raw_b)])
        except ValueError:
            errors.append(
                f"Skor {segment_label} {index} harus berupa bilangan bulat."
            )
    return games, errors


def _parse_capsa_score_form():
    """Parse per-round remaining-card counts into ([points_a, points_b], ...)
    round segments plus the raw per-player breakdown for crosscheck display."""
    from domain.scoring.capsa_susun import round_points, validate_round_hand, ROUNDS_PER_MATCH

    segments = []
    breakdown = []
    errors = []
    gap_found = False
    for index in range(1, ROUNDS_PER_MATCH + 1):
        raw = {
            key: request.form.get(f"round{index}_{key}", "").strip()
            for key in ("a1", "a2", "b1", "b2")
        }
        if not any(raw.values()):
            gap_found = True
            continue
        if not all(raw.values()):
            errors.append(f"Ronde {index}: sisa kartu keempat pemain harus diisi semua.")
            continue
        if gap_found:
            errors.append("Ronde harus diisi berurutan tanpa melewati ronde kosong.")
        try:
            cards = {key: int(value) for key, value in raw.items()}
        except ValueError:
            errors.append(f"Ronde {index}: sisa kartu harus berupa bilangan bulat.")
            continue
        hand_errors = validate_round_hand(cards["a1"], cards["a2"], cards["b1"], cards["b2"])
        if hand_errors:
            errors.extend(f"Ronde {index}: {msg}" for msg in hand_errors)
            continue
        twos = {}
        for key in ("a1", "a2", "b1", "b2"):
            raw_twos = request.form.get(f"round{index}_{key}_twos", "").strip()
            try:
                twos[key] = int(raw_twos) if raw_twos else 0
            except ValueError:
                twos[key] = 0
        points_a, points_b = round_points(
            cards["a1"], cards["a2"], cards["b1"], cards["b2"],
            twos["a1"], twos["a2"], twos["b1"], twos["b2"],
        )
        segments.append([points_a, points_b])
        breakdown.append({**cards, "points_a": points_a, "points_b": points_b})
    return segments, breakdown, errors


def data_context():
    teams = utils.load_teams()
    matches = utils.load_matches()
    config = utils.load_config()
    return teams, matches, config


def _selected_sport():
    requested = request.args.get("sport", "").strip()
    known = {sport["key"] for sport in utils.list_sports()}
    return requested if requested in known else "all"


def _filter_matches_by_selected_sport(matches):
    selected = _selected_sport()
    if selected == "all":
        return matches
    return [
        match for match in matches
        if match.get("sport_key", "table-tennis") == selected
    ]


def _sport_url(endpoint, **values):
    selected = _selected_sport()
    if selected != "all" and "sport" not in values:
        values["sport"] = selected
    return url_for(endpoint, **values)


def _switch_sport_url(sport_key):
    values = dict(request.view_args or {})
    values.update(request.args.to_dict(flat=True))
    if sport_key == "all":
        values.pop("sport", None)
    else:
        values["sport"] = sport_key
    return url_for(request.endpoint or "index", **values)


@app.context_processor
def inject_globals():
    config = utils.load_config()
    sports = utils.list_sports()
    selected_sport = _selected_sport()
    return {
        "config": config,
        "tournament_date_range": utils.format_date_range_id(config["start_date"], config["end_date"]),
        "sports": sports,
        "selected_sport": selected_sport,
        "sport_url": _sport_url,
        "switch_sport_url": _switch_sport_url,
        "storage_backend": utils.STORAGE_BACKEND,
        "is_admin": bool(session.get("is_admin")),
        "admin_role": (session.get("admin_role") or "general") if session.get("is_admin") else None,
        "admin_sport": session.get("admin_sport"),
        "admin_label": session.get("admin_label"),
        "now": utils.now_wib(config.get("timezone")),
        "csrf_token": get_csrf_token(),
        "license": license_client.get_license(),
    }


@app.teardown_appcontext
def _close_db_connection(exception=None):
    utils.close_request_connection()


# ---------- public routes ----------

@app.route("/")
def index():
    teams, matches, config = data_context()
    all_matches = matches
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]
    today = utils.now_wib(config.get("timezone")).strftime("%Y-%m-%d")
    selected_sport = _selected_sport()
    competition = competition_context(
        matches, teams, config, sport_key=selected_sport
    )

    upcoming_candidates = [
        m for m in enriched if m["status"] in ("scheduled", "live") and m["date"] >= today
    ]
    if selected_sport == "all":
        # Chronological-only would let whichever sport starts earliest (e.g. Padel
        # on day 1) fill every slot before Badminton/Tenis Meja ever get a look in,
        # so cap how many of the nearest matches each sport can contribute.
        by_sport = {}
        for m in sorted(upcoming_candidates, key=utils.match_sort_key):
            by_sport.setdefault(m.get("sport_key", "table-tennis"), []).append(m)
        upcoming = sorted(
            [m for sport_matches in by_sport.values() for m in sport_matches[:2]],
            key=utils.match_sort_key,
        )
    else:
        upcoming = sorted(upcoming_candidates, key=utils.match_sort_key)[:6]
    live_now = [m for m in enriched if m["status"] == "live"]
    recent = sorted(
        [m for m in enriched if m["status"] == "completed"],
        key=lambda m: (m["date"], m["time"]), reverse=True,
    )[:4]

    total_matches = len(matches)
    completed_count = sum(1 for m in matches if m["status"] == "completed")
    incomplete_count = sum(
        1 for m in matches if m["status"] not in ("completed", "cancelled")
    )

    all_comments = []
    for m in enriched:
        if m["status"] == "completed":
            continue
        if m.get("comments"):
            for c in m["comments"]:
                all_comments.append({
                    "name": c["name"],
                    "comment": c["comment"],
                    "match_label": f"{m['team_a_code']} vs {m['team_b_code']}",
                    "at": c["at"]
                })
    all_comments.sort(key=lambda x: x["at"], reverse=True)
    recent_comments = all_comments[:15]
    
    next_match = None
    for m in upcoming:
        if m["status"] == "scheduled" and m.get("time") and m.get("team_a") and m.get("team_b"):
            next_match = m
            break
            
    is_live = len(live_now) > 0
    tournament_ended = (
        config.get("status") == "completed"
        and all(
            match.get("status") in {"completed", "cancelled"}
            for match in all_matches
        )
    )

    champion_divisions = []
    hero_gallery_photos = []
    if tournament_ended:
        champion_divisions = [
            division for division in competition.get("divisions", [])
            if division.get("champion_team")
        ]

        doc_photos = []
        for m in enriched:
            for doc in m.get("docs", []):
                doc_photos.append({"url": doc["url"], "uploaded_at": doc.get("uploaded_at", "")})
        doc_photos.sort(key=lambda p: p["uploaded_at"], reverse=True)
        hero_gallery_photos = doc_photos[:16]

    full_structure = utils.load_competition_structure()
    sport_cards = []
    for sport in utils.list_sports():
        sport_matches = [
            match for match in all_matches
            if match.get("sport_key", "table-tennis") == sport["key"]
        ]
        sport_enriched = [enrich_match(match, teams) for match in sport_matches]
        sport_upcoming = sorted(
            [
                match for match in sport_enriched
                if match.get("status") in {"scheduled", "live"}
                and (match.get("date") or "") >= today
            ],
            key=utils.match_sort_key,
        )
        sport_cards.append(
            {
                **sport,
                "match_count": len(sport_matches),
                "live_count": sum(
                    1 for match in sport_matches if match.get("status") == "live"
                ),
                "completed_count": sum(
                    1 for match in sport_matches
                    if match.get("status") == "completed"
                ),
                "next_match": sport_upcoming[0] if sport_upcoming else None,
                "division_count": sum(
                    1 for division in full_structure.get("divisions", [])
                    if division.get("sport_key") == sport["key"]
                    and division.get("enabled", True)
                ),
            }
        )

    visible_team_codes = {
        code for match in matches
        for code in (match.get("team_a"), match.get("team_b")) if code
    }

    return render_template(
        "index.html",
        upcoming=upcoming, live_now=live_now, recent=recent,
        total_matches=total_matches, completed_count=completed_count,
        incomplete_count=incomplete_count,
        team_count=len(visible_team_codes), division_count=len(competition["divisions"]),
        recent_comments=recent_comments,
        next_match=next_match, is_live=is_live,
        tournament_ended=tournament_ended,
        champion_divisions=champion_divisions,
        sport_cards=sport_cards,
        hero_gallery_photos=hero_gallery_photos,
    )


@app.route("/jadwal")
def jadwal():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]

    kategori = request.args.get("kategori", "")
    grup = request.args.get("grup", "")
    tanggal = request.args.get("tanggal", "")
    status = request.args.get("status", "")
    q = request.args.get("q", "").strip()

    filtered = enriched
    if kategori:
        filtered = [m for m in filtered if m["category"] == kategori]
    if grup:
        filtered = [m for m in filtered if m["group"] == grup]
    if tanggal:
        filtered = [m for m in filtered if m["date"] == tanggal]
    if status:
        filtered = [m for m in filtered if m["status"] == status]
    if q:
        needle = q.lower()
        filtered = [
            m for m in filtered
            if needle in m["team_a_code"].lower()
            or needle in m["team_b_code"].lower()
            or needle in m["team_a_player1"].lower()
            or needle in m["team_a_player2"].lower()
            or needle in m["team_b_player1"].lower()
            or needle in m["team_b_player2"].lower()
        ]

    filtered.sort(key=utils.match_sort_key)
    dates = sorted({m["date"] for m in enriched})
    categories = [
        category for category in config.get("categories", [])
        if _selected_sport() == "all"
        or category.get("sport_key", "table-tennis") == _selected_sport()
    ]
    groups = sorted({m["group"] for m in enriched if m.get("group")})

    status_counts = Counter(m["status"] for m in enriched)
    schedule_stats = {
        "total": len(enriched),
        "scheduled": status_counts.get("scheduled", 0),
        "live": status_counts.get("live", 0),
        "completed": status_counts.get("completed", 0),
        "postponed": status_counts.get("postponed", 0),
    }
    schedule_stats["progress"] = round(
        (schedule_stats["completed"] / schedule_stats["total"]) * 100
    ) if schedule_stats["total"] else 0

    schedule_days = []
    for match in filtered:
        if not schedule_days or schedule_days[-1]["date"] != match["date"]:
            schedule_days.append({
                "date": match["date"],
                "date_label": match["date_label"],
                "matches": [],
            })
        schedule_days[-1]["matches"].append(match)

    return render_template(
        "jadwal.html", matches=filtered, schedule_days=schedule_days,
        schedule_stats=schedule_stats,
        categories=categories, groups=groups, dates=dates,
        kategori=kategori, grup=grup, tanggal=tanggal, status=status, q=q,
        active_filter_count=sum(bool(value) for value in (kategori, grup, tanggal, status, q)),
    )


@app.route("/kalender")
def kalender():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]

    by_date = {}
    final_dates = set()
    for m in enriched:
        by_date.setdefault(m["date"], []).append(m)
        if m.get("stage_type") == "final":
            final_dates.add(m["date"])
    for d in by_date:
        by_date[d].sort(key=utils.match_sort_key)

    start = datetime.strptime(config["start_date"], "%Y-%m-%d")
    end = datetime.strptime(config["end_date"], "%Y-%m-%d")
    days = []
    cur = start
    while cur <= end:
        key = cur.strftime("%Y-%m-%d")
        days.append({
            "date": key,
            "day_num": cur.day,
            "day_name": utils.DAY_NAMES[cur.weekday()],
            "matches": by_date.get(key, []),
            "is_buffer": key in config.get("buffer_dates", []),
            "is_weekend": cur.weekday() >= 5,
            "is_final": key in final_dates,
            "is_closing": key == config.get("closing_date"),
        })
        cur += timedelta(days=1)

    return render_template("kalender.html", days=days, month_label=f"{utils.MONTH_NAMES[start.month]} {start.year}")


@app.route("/kalender/export")
def kalender_export():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]
    enriched.sort(key=utils.match_sort_key)

    sports_by_key = {sport["key"]: sport for sport in utils.list_sports()}
    categories_by_key = {cat["key"]: cat.get("label", cat["key"]) for cat in config.get("categories", [])}
    selected_sport = _selected_sport()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rundown Jadwal"

    headers = [
        "No", "Tanggal", "Waktu", "Cabor", "Kategori", "Grup",
        "Tim A", "Tim B", "Skor", "Status", "Lokasi",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for idx, m in enumerate(enriched, start=1):
        sport_name = sports_by_key.get(m.get("sport_key", "table-tennis"), {}).get("name", m.get("sport_key", ""))
        kategori_label = categories_by_key.get(m.get("category"), m.get("category", ""))
        if m["status"] in ("completed", "live"):
            skor = f"{m['sets_a']} - {m['sets_b']}"
        else:
            skor = "-"
        ws.append([
            idx,
            m.get("date_label") or m.get("date", ""),
            m.get("time", ""),
            sport_name,
            kategori_label,
            "Final" if m.get("group") == "FINAL" else m.get("group", ""),
            m.get("team_a_code", ""),
            m.get("team_b_code", ""),
            skor,
            m.get("status_label", ""),
            m.get("court", ""),
        ])

    widths = [5, 16, 8, 14, 26, 8, 14, 14, 8, 12, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    sport_label = sports_by_key.get(selected_sport, {}).get("name", "Semua Cabor") if selected_sport != "all" else "Semua Cabor"
    filename = f"Rundown Jadwal {sport_label} - {config.get('tournament_short_name', 'Turnamen')}.xlsx"
    filename = filename.replace("/", "-")

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/aturan")
def aturan():
    structure = utils.load_competition_structure()
    selected = _selected_sport()
    sports_by_key = {sport["key"]: sport for sport in utils.list_sports()}
    rule_sports = []
    for sport_key, profiles in structure.get("rule_profiles_by_sport", {}).items():
        if selected != "all" and sport_key != selected:
            continue
        sport = sports_by_key.get(
            sport_key,
            {"key": sport_key, "name": sport_key, "icon": "🏟️", "enabled": False},
        )
        rule_sports.append({**sport, "profiles": profiles})
    divisions = [
        division for division in structure.get("divisions", [])
        if selected == "all" or division.get("sport_key") == selected
    ]
    return render_template(
        "aturan.html", rule_sports=rule_sports, rule_divisions=divisions,
        capsa_participants=CAPSA_PARTICIPANTS,
    )


@app.route("/tim-peserta")
def tim_peserta():
    teams = utils.load_teams()
    config = utils.load_config()
    sites = config.get("sites", [])
    categories = config.get("categories", [])
    sports = [sport for sport in utils.list_sports() if sport.get("enabled")]
    selected = _selected_sport()

    teams_by_site = {}
    for team_id, team in teams.items():
        teams_by_site.setdefault(team.get("site_code"), []).append({**team, "id": team_id})

    site_cards = []
    for site in sites:
        site_teams = teams_by_site.get(site["code"], [])
        sport_sections = []
        for sport in sports:
            if selected != "all" and sport["key"] != selected:
                continue
            category_groups = []
            for cat in categories:
                if cat.get("sport_key") != sport["key"]:
                    continue
                cat_teams = sorted(
                    (t for t in site_teams if t.get("category") == cat["key"]),
                    key=lambda t: (t.get("group") or "", t.get("id") or ""),
                )
                if cat_teams:
                    category_groups.append({
                        "label": cat.get("label", cat["key"]),
                        "entrant_type": cat.get("entrant_type", "pair"),
                        "teams": cat_teams,
                    })
            if category_groups:
                sport_sections.append({
                    "key": sport["key"],
                    "name": sport["name"],
                    "icon": sport["icon"],
                    "category_groups": category_groups,
                })
        if sport_sections:
            site_cards.append({
                "code": site["code"],
                "name": site["name"],
                "team_count": sum(
                    len(group["teams"])
                    for section in sport_sections
                    for group in section["category_groups"]
                ),
                "sport_count": len(sport_sections),
                "sport_sections": sport_sections,
            })

    participant_ids = set()
    for site_teams in teams_by_site.values():
        for t in site_teams:
            for key in ("participant_id1", "participant_id2", "reserve_participant_id"):
                if t.get(key):
                    participant_ids.add(t[key])
            for reserve in t.get("reserves") or []:
                if reserve.get("participant_id"):
                    participant_ids.add(reserve["participant_id"])

    stats = {
        "site_count": len(site_cards),
        "team_count": sum(site["team_count"] for site in site_cards),
        "participant_count": len(participant_ids),
    }

    return render_template("tim_peserta.html", site_cards=site_cards, stats=stats)


@app.route("/klasemen")
def klasemen():
    teams, matches, config = data_context()
    selected = _selected_sport()
    filtered_matches = _filter_matches_by_selected_sport(matches)
    competition = competition_context(
        filtered_matches, teams, config, sport_key=selected
    )
    medal_tally = (
        utils.compute_medal_tally(competition.get("divisions", []), teams, config)
        if selected == "all" else None
    )
    return render_template(
        "klasemen.html", competition=competition, medal_tally=medal_tally
    )


@app.route("/admin/juara/<category>/<group>", methods=["POST"])
@login_required
def admin_upload_champion_photo(category, group):
    teams, matches, config = data_context()
    redirect_to = request.form.get("redirect_to", "klasemen")
    redirect_endpoint = "bracket" if redirect_to == "bracket" else "klasemen"

    competition = competition_context(matches, teams, config)
    division = next(
        (
            item for item in competition.get("divisions", [])
            if item.get("key") == category
        ),
        None,
    )
    if division:
        _assert_sport_access(division.get("sport_key"))
    champion = None
    if division:
        if division.get("champion_target") == group:
            champion = division.get("champion")
        if not champion:
            champion = next(
                (
                    group_view.get("champion")
                    for stage in division.get("stages_view", [])
                    for group_view in stage.get("groups_view", [])
                    if group_view.get("key") == group
                ),
                None,
            )
    if not champion:
        flash("Kategori/grup ini belum punya juara (belum semua pertandingan selesai).", "error")
        return redirect(url_for(redirect_endpoint))

    action = request.form.get("action")
    champions = config.setdefault("champions", {})
    key = f"{category}_{group}"

    if action == "delete_champion_photo":
        old = champions.get(key, {}).get("photo_url")
        if old:
            delete_from_supabase(old)
            utils.set_champion_photo(key)
            flash("Foto juara berhasil dihapus.", "success")
        return redirect(url_for(redirect_endpoint))

    file = request.files.get("champion_photo_cam") or request.files.get("champion_photo")
    if not file or not file.filename:
        flash("Pilih file foto juara terlebih dahulu.", "error")
        return redirect(url_for(redirect_endpoint))

    file_url, error = compress_and_upload_image(file, f"champion_{key}")
    if error:
        flash(f"Gagal mengunggah foto juara: {error}", "error")
        return redirect(url_for(redirect_endpoint))

    old = champions.get(key, {}).get("photo_url")
    if old:
        delete_from_supabase(old)

    utils.set_champion_photo(
        key,
        file_url,
        utils.now_wib(config.get("timezone")).isoformat(timespec="seconds"),
    )
    flash("Foto juara berhasil diunggah.", "success")
    return redirect(url_for(redirect_endpoint))


@app.route("/bracket")
def bracket():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    competition = competition_context(
        matches, teams, config, sport_key=_selected_sport()
    )
    return render_template("bracket.html", competition=competition)


@app.route("/live")
def live():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]
    live_now = [m for m in enriched if m["status"] == "live"]
    today = utils.now_wib(config.get("timezone")).strftime("%Y-%m-%d")
    today_matches = sorted(
        [m for m in enriched if m["date"] == today],
        key=lambda m: (m["time"] == "", m["time"], m["court"]),
    )
    youtube_video_id = utils.extract_youtube_video_id(config.get("youtube_embed_url", ""))
    embed_domain = request.host.split(":")[0]
    return render_template(
        "live.html", live_now=live_now, today_matches=today_matches,
        youtube_video_id=youtube_video_id, embed_domain=embed_domain,
        live_chat_messages=utils.list_live_chat_messages(),
    )


@app.route("/api/v1/live-chat")
def api_v1_live_chat():
    after_id = request.args.get("after", "").strip() or None
    return jsonify({"data": utils.list_live_chat_messages(after_id=after_id)})


@app.route("/live-chat/send", methods=["POST"])
def send_live_chat_message():
    name = utils.censor_text(request.form.get("name", "").strip()[:24])
    message = utils.censor_text(request.form.get("message", "").strip()[:120])
    if not name or not message:
        return jsonify({"ok": False, "error": "Nama dan pesan wajib diisi."}), 400
    if _rate_limited(f"chat:{request.remote_addr or 'unknown'}", 8, 30):
        return jsonify({"ok": False, "error": "Terlalu banyak pesan, coba lagi sebentar."}), 429
    entry = utils.add_live_chat_message(name, message)
    return jsonify({"ok": True, "message": entry})


@app.route("/admin/live-chat/delete", methods=["POST"])
@login_required
def admin_delete_live_chat_message():
    message_id = request.form.get("message_id", "").strip()
    if message_id and utils.delete_live_chat_message(message_id):
        flash("Pesan chat berhasil dihapus.", "success")
    else:
        flash("Pesan chat tidak ditemukan.", "error")
    return redirect(url_for("live"))


@app.route("/rekap")
def rekap():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]
    completed = sorted(
        [m for m in enriched if m["status"] == "completed"],
        key=lambda m: (m["date"], m["time"]), reverse=True,
    )
    kategori = request.args.get("kategori", "")
    if kategori:
        completed = [m for m in completed if m["category"] == kategori]
    categories = [
        category for category in config.get("categories", [])
        if _selected_sport() == "all"
        or category.get("sport_key", "table-tennis") == _selected_sport()
    ]
    return render_template(
        "rekap.html", matches=completed, kategori=kategori,
        categories=categories,
    )


@app.route("/galeri")
def galeri():
    teams, matches, config = data_context()
    matches = _filter_matches_by_selected_sport(matches)
    enriched = [enrich_match(m, teams) for m in matches]
    photos = []
    for m in enriched:
        if m.get("docs"):
            for doc in m["docs"]:
                photos.append({
                    "url": doc["url"],
                    "uploaded_at": doc.get("uploaded_at", ""),
                    "match_id": m["id"],
                    "team_a_code": m["team_a_code"],
                    "team_b_code": m["team_b_code"],
                    "category_label": m["category_label"],
                    "round_label": m["round_label"],
                    "date_label": m["date_label"],
                    "time": m["time"],
                    "is_standalone": False,
                })
    for gp in utils.list_gallery_photos():
        uploaded_at = gp.get("uploaded_at", "")
        try:
            date_label = utils.format_date_id(uploaded_at.split("T")[0])
        except (ValueError, IndexError):
            date_label = ""
        photos.append({
            "url": gp["url"],
            "uploaded_at": uploaded_at,
            "photo_id": gp.get("id"),
            "category_label": "Galeri",
            "date_label": date_label,
            "is_standalone": True,
        })
    photos.sort(key=lambda p: p["uploaded_at"], reverse=True)
    return render_template("galeri.html", photos=photos)


@app.route("/admin/galeri/upload", methods=["POST"])
@login_required
def admin_upload_gallery_photos():
    files = [f for f in request.files.getlist("gallery_photos") if f and f.filename]
    if not files:
        flash("Pilih minimal satu foto untuk diunggah.", "error")
        return redirect(url_for("galeri"))
    if len(files) > GALLERY_UPLOAD_BATCH_LIMIT:
        flash(f"Maksimal {GALLERY_UPLOAD_BATCH_LIMIT} foto per sekali unggah.", "error")
        return redirect(url_for("galeri"))

    uploaded_urls = []
    upload_errors = []
    for file in files:
        file_url, error = compress_and_upload_image(file, "gallery")
        if error:
            upload_errors.append(f"{file.filename}: {error}")
        elif file_url:
            uploaded_urls.append(file_url)

    if uploaded_urls:
        utils.add_gallery_photos(uploaded_urls)
        flash(f"{len(uploaded_urls)} foto berhasil diunggah dan dikompres ke galeri.", "success")
    if upload_errors:
        flash("Sebagian foto gagal diunggah: " + "; ".join(upload_errors), "error")
    return redirect(url_for("galeri"))


@app.route("/admin/galeri/delete", methods=["POST"])
@login_required
def admin_delete_gallery_photo():
    photo_id = request.form.get("photo_id", "").strip()
    url = utils.delete_gallery_photo(photo_id) if photo_id else None
    if url:
        delete_from_supabase(url)
        flash("Foto galeri berhasil dihapus.", "success")
    else:
        flash("Foto galeri tidak ditemukan.", "error")
    return redirect(url_for("galeri"))


def _match_detail_context(match_id, is_modal):
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        abort(404)
    enriched = enrich_match(m, teams)
    ctx = {
        "m": enriched,
        "share_text": utils.build_share_text(enriched),
        "is_modal": is_modal,
    }
    if session.get("is_admin"):
        ctx["raw"] = m
        ctx["teams"] = teams
        ctx.update(qualification_context(m, matches, teams, config))
    return ctx


@app.route("/pertandingan/<match_id>")
def match_detail(match_id):
    return render_template("match_detail.html", **_match_detail_context(match_id, is_modal=False))


@app.route("/pertandingan/<match_id>/fragment")
def match_fragment(match_id):
    return render_template("_match_detail_content.html", **_match_detail_context(match_id, is_modal=True))


@app.route("/pertandingan/<match_id>/komentar", methods=["POST"])
def add_comment(match_id):
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        abort(404)

    sport = request.args.get("sport")
    target = url_for(
        "match_fragment", match_id=match_id, sport=sport
    ) if request.form.get("modal") else url_for(
        "match_detail", match_id=match_id, sport=sport
    ) + "#komentar"

    session_key = f"comments_{match_id}"
    comments_made = session.get(session_key, 0)
    if comments_made >= 3:
        flash("Batas maksimum 3 komentar per sesi.", "error")
        return redirect(target)

    name = request.form.get("name", "").strip()
    comment = request.form.get("comment", "").strip()
    if not name or not comment:
        flash("Nama dan komentar wajib diisi.", "error")
        return redirect(target)
        
    if len(comment) > 200:
        flash("Komentar terlalu panjang (maksimal 200 karakter).", "error")
        return redirect(target)

    # Censor bad words
    clean_name = utils.censor_text(name[:60])
    clean_comment = utils.censor_text(comment[:200])

    new_comment = {
        "name": clean_name,
        "comment": clean_comment,
        "at": utils.now_wib().isoformat(timespec="seconds"),
    }

    def append_comment(current):
        if current.get("status") == "completed":
            raise ValueError("Komentar ditutup setelah pertandingan selesai.")
        current.setdefault("comments", []).append(new_comment)

    try:
        utils.update_match(match_id, append_comment)
    except (utils.MatchNotFoundError, ValueError) as exc:
        flash(str(exc), "error")
        return redirect(target)

    session[session_key] = comments_made + 1
    flash("Komentar terkirim. Panitia akan meninjau permintaan Anda.", "success")
    return redirect(target)


@app.route("/pertandingan/<match_id>/vote", methods=["POST"])
def vote_match(match_id):
    team = request.form.get("team")
    emoji = request.form.get("emoji")
    
    valid_emojis = ["🔥", "👏", "❤️", "😲", "💪"]
    if emoji not in valid_emojis or team not in ("a", "b"):
        return jsonify({"success": False, "message": "Data tidak valid"}), 400
        
    session_key = f"vote_{match_id}"
    vote_count = session.get(session_key, 0)
    
    if vote_count >= 2:
        return jsonify({"success": False, "message": "Batas maksimum 2 vote per sesi."}), 403
        
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        return jsonify({"success": False, "message": "Pertandingan tidak ditemukan"}), 404
        
    # Check time limit
    try:
        match_dt_str = f"{m['date']} {m['time']}"
        match_dt = datetime.strptime(match_dt_str, "%Y-%m-%d %H:%M").replace(
            tzinfo=ZoneInfo(config.get("timezone", "Asia/Jakarta"))
        )
        if utils.now_wib(config.get("timezone")) >= match_dt - timedelta(hours=1):
            return jsonify({"success": False, "message": "Voting ditutup 1 jam sebelum pertandingan."}), 403
    except:
        pass
        
    def add_vote(current):
        if "votes" not in current:
            current["votes"] = {
                "a": {value: 0 for value in valid_emojis},
                "b": {value: 0 for value in valid_emojis},
            }
        current["votes"].setdefault(team, {})
        current["votes"][team].setdefault(emoji, 0)
        current["votes"][team][emoji] += 1

    try:
        m = utils.update_match(match_id, add_vote)
    except utils.MatchNotFoundError:
        return jsonify({"success": False, "message": "Pertandingan tidak ditemukan"}), 404

    session[session_key] = vote_count + 1
    
    total_a = sum(m["votes"]["a"].values())
    total_b = sum(m["votes"]["b"].values())
    total_votes = total_a + total_b
    pct_a = int(round((total_a / total_votes) * 100)) if total_votes > 0 else 0
    pct_b = 100 - pct_a if total_votes > 0 else 0
    
    return jsonify({
        "success": True,
        "new_count": m["votes"][team][emoji],
        "votes_left": 2 - (vote_count + 1),
        "pct_a": pct_a,
        "pct_b": pct_b
    })



# ---------- Share card (Open Graph) ----------

@app.route("/og/match/<match_id>.png")
def og_match_image(match_id):
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        abort(404)
    buf = og_image.generate_match_card(enrich_match(m, teams), config["tournament_short_name"])
    resp = send_file(buf, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=300"
    return resp


@app.route("/og/default.png")
def og_default_image():
    config = utils.load_config()
    title = request.args.get("title", config["tournament_short_name"])
    subtitle = request.args.get(
        "subtitle", f'{config["start_date"]} – {config["end_date"]} · {config["venue"]}'
    )
    tag = request.args.get("tag", "ROUND ROBIN")
    buf = og_image.generate_default_card(title, subtitle, tag=tag)
    resp = send_file(buf, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


# ---------- JSON API (buat live polling) ----------

@app.route("/api/matches")
def api_matches():
    teams, matches, config = data_context()
    return jsonify([enrich_match(m, teams) for m in matches])


@app.route("/api/matches/<match_id>")
def api_match(match_id):
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        return jsonify({"error": "not found"}), 404
    return jsonify(enrich_match(m, teams))


def _api_error(code, message, status=400, fields=None):
    payload = {"error": {"code": code, "message": message}}
    if fields:
        payload["error"]["fields"] = fields
    return jsonify(payload), status


def _api_match_document(match, teams, timezone_name, detail=False):
    team_a = teams.get(match.get("team_a"), {})
    team_b = teams.get(match.get("team_b"), {})
    winner = teams.get(match.get("winner"), {})
    _, score_errors = utils.validate_recorded_result(match)
    if match.get("sport_key") == "capsa-susun":
        segments_a = sum(s[0] for s in match.get("sets") or [])
        segments_b = sum(s[1] for s in match.get("sets") or [])
    else:
        segments_a, segments_b = utils.compute_sets_won(match.get("sets") or [])
    document = {
        "id": match["id"],
        "sport": {"key": match.get("sport_key", "table-tennis")},
        "division": {
            "key": match.get("category"),
            "name": match.get("category_label"),
        },
        "stage": {
            "type": match.get("stage_type") or (
                "final" if match.get("group") == "FINAL" else "group"
            ),
            "name": match.get("stage_label") or match.get("round_label"),
            "group": match.get("group"),
            "round": match.get("round"),
            "round_label": match.get("round_label"),
        },
        "entrants": {
            "a": {
                "code": team_a.get("code", match.get("team_a")),
                "members": [
                    name for name in (team_a.get("player1"), team_a.get("player2"))
                    if name
                ],
            },
            "b": {
                "code": team_b.get("code", match.get("team_b")),
                "members": [
                    name for name in (team_b.get("player1"), team_b.get("player2"))
                    if name
                ],
            },
        },
        "schedule": {
            "date": match.get("date") or None,
            "time": match.get("time") or None,
            "timezone": timezone_name,
            "court": match.get("court") or None,
        },
        "status": match.get("status"),
        "status_label": STATUS_LABELS.get(match.get("status"), match.get("status")),
        "score": {
            "segments": match.get("sets") or [],
            "segments_won": {"a": segments_a, "b": segments_b},
            "segment_term": match.get("_segment_term", "game"),
            "winner": winner.get("code", match.get("winner")),
            "result_type": "walkover" if match.get("walkover") else "normal",
            "valid": bool(match.get("_result_valid", True)) and not score_errors,
        },
        "version": int(match.get("version", 0)),
        "links": {"self": url_for("api_v1_match", match_id=match["id"], _external=True)},
    }
    if detail:
        document.update(
            {
                "notes": match.get("notes", ""),
                "comments": [
                    {
                        "author": comment.get("name"),
                        "body": comment.get("comment"),
                        "created_at": comment.get("at"),
                    }
                    for comment in match.get("comments", [])
                ],
                "reactions": match.get("votes", {"a": {}, "b": {}}),
                "media": [
                    {
                        "url": media.get("url"),
                        "uploaded_at": media.get("uploaded_at"),
                    }
                    for media in match.get("docs", [])
                ],
            }
        )
    return document


def _json_with_etag(payload, status=200):
    canonical = json.dumps(
        payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False
    ).encode("utf-8")
    etag = hashlib.sha256(canonical).hexdigest()
    response = jsonify(payload)
    response.status_code = status
    response.set_etag(etag)
    if request.if_none_match and request.if_none_match.contains(etag):
        response.status_code = 304
        response.set_data(b"")
    return response


@app.route("/api/v1/sports")
def api_v1_sports():
    payload = {
        "data": utils.list_sports(),
        "meta": {"storage_backend": utils.STORAGE_BACKEND},
    }
    return _json_with_etag(payload)


@app.route("/api/v1/matches")
def api_v1_matches():
    try:
        limit = int(request.args.get("limit", "25"))
    except ValueError:
        return _api_error(
            "invalid_request", "limit must be an integer.",
            fields={"limit": "invalid"},
        )
    if not 1 <= limit <= 100:
        return _api_error(
            "invalid_request", "limit must be between 1 and 100.",
            fields={"limit": "out_of_range"},
        )

    status = request.args.get("status") or None
    valid_statuses = {
        "draft", "scheduled", "check_in", "live", "postponed",
        "suspended", "cancelled", "completed",
    }
    if status and status not in valid_statuses:
        return _api_error(
            "invalid_request", "Unknown match status.",
            fields={"status": "invalid"},
        )
    sport = request.args.get("sport") or None
    known_sports = {item["key"] for item in utils.list_sports()}
    if sport and sport not in known_sports:
        return _api_error(
            "invalid_request", "Unknown sport.", fields={"sport": "invalid"}
        )
    match_date = request.args.get("date") or None
    if match_date:
        try:
            datetime.strptime(match_date, "%Y-%m-%d")
        except ValueError:
            return _api_error(
                "invalid_request", "date must use YYYY-MM-DD.",
                fields={"date": "invalid"},
            )

    try:
        matches, next_cursor = utils.list_api_matches(
            sport=sport,
            division=request.args.get("division") or None,
            status=status,
            match_date=match_date,
            limit=limit,
            cursor_value=request.args.get("cursor") or None,
        )
    except ValueError as exc:
        return _api_error(
            "invalid_cursor", str(exc), fields={"cursor": "invalid"}
        )

    teams = utils.load_teams()
    timezone_name = utils.load_config().get("timezone", "Asia/Jakarta")
    data = [
        _api_match_document(match, teams, timezone_name)
        for match in matches
    ]
    next_url = None
    if next_cursor:
        query = request.args.to_dict(flat=True)
        query["cursor"] = next_cursor
        next_url = url_for("api_v1_matches", _external=True, **query)
    payload = {
        "data": data,
        "meta": {
            "count": len(data), "limit": limit,
            "next_cursor": next_cursor,
        },
        "links": {"self": request.url, "next": next_url},
    }
    return _json_with_etag(payload)


@app.route("/api/v1/matches/<match_id>")
def api_v1_match(match_id):
    match = utils.get_api_match(match_id)
    if not match:
        return _api_error("not_found", "Match not found.", status=404)
    payload = {
        "data": _api_match_document(
            match,
            utils.load_teams(),
            utils.load_config().get("timezone", "Asia/Jakarta"),
            detail=True,
        )
    }
    return _json_with_etag(payload)


@app.route("/api/v1/standings")
def api_v1_standings():
    sport = request.args.get("sport") or None
    known_sports = {item["key"] for item in utils.list_sports()}
    if sport and sport not in known_sports:
        return _api_error(
            "invalid_request", "Unknown sport.", fields={"sport": "invalid"}
        )
    teams, matches, config = data_context()
    competition = competition_context(
        matches, teams, config, sport_key=sport or "all"
    )
    division_filter = request.args.get("division") or None
    stage_filter = request.args.get("stage") or None
    group_filter = request.args.get("group") or None
    data = []
    for division in competition.get("divisions", []):
        if division_filter and division["key"] != division_filter:
            continue
        for stage in division.get("stages_view", []):
            if stage.get("type") not in {"group", "round_robin"}:
                continue
            if stage_filter and stage.get("key") != stage_filter:
                continue
            for group in stage.get("groups_view", []):
                if group_filter and group.get("key") != group_filter:
                    continue
                data.append(
                    {
                        "sport": division["sport_key"],
                        "division": {
                            "key": division["key"], "name": division["name"]
                        },
                        "stage": {
                            "key": stage["key"], "name": stage["name"],
                            "type": stage["type"],
                        },
                        "group": {"key": group["key"], "name": group["name"]},
                        "complete": group["complete"],
                        "champion": utils.team_short(
                            teams, group.get("champion")
                        ) if group.get("champion") else None,
                        "policy": division.get("standing_policy"),
                        "rows": [
                            {
                                "rank": row["rank"],
                                "entrant": {
                                    "code": row["display_code"],
                                    "members": [
                                        name for name in (
                                            row.get("player1"), row.get("player2")
                                        ) if name
                                    ],
                                },
                                "played": row["played"], "wins": row["win"],
                                "losses": row["loss"], "points": row["points"],
                                "segments_for": row["set_win"],
                                "segments_against": row["set_loss"],
                                "segment_difference": row["set_diff"],
                                "points_for": row["point_win"],
                                "points_against": row["point_loss"],
                                "point_difference": row["point_diff"],
                                "tie_break": row["tie_break"],
                            }
                            for row in group.get("rows", [])
                        ],
                    }
                )
    return _json_with_etag({"data": data, "meta": {"count": len(data)}})


# ---------- admin ----------

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    accounts = _configured_accounts()
    if request.method == "POST":
        client_key = _login_client_key()
        remaining = _login_lock_remaining(client_key)
        account_key = request.form.get("account_key", "")
        if remaining:
            flash(
                f"Terlalu banyak percobaan login. Coba lagi dalam {remaining} detik.",
                "error",
            )
            return render_template("admin/login.html", accounts=accounts, selected_account_key=account_key), 429

        if not accounts:
            flash(
                "Login admin belum dikonfigurasi. Set ADMIN_PASSWORD_HASH di environment.",
                "error",
            )
            return render_template("admin/login.html", accounts=accounts, selected_account_key=account_key), 503

        password = request.form.get("password", "")
        account = next((a for a in accounts if a["key"] == account_key), None)

        if account and _verify_account_password(account, password):
            _clear_login_failures(client_key)
            session.clear()
            session["is_admin"] = True
            session["admin_key"] = account["key"]
            session["admin_role"] = account["role"]
            session["admin_sport"] = account["sport_key"]
            session["admin_label"] = account["label"]
            session[CSRF_SESSION_KEY] = secrets.token_urlsafe(32)
            session.permanent = True
            if account["role"] == "gallery":
                return redirect(url_for("galeri"))
            nxt = _safe_internal_next(request.args.get("next")) or url_for("admin_dashboard")
            return redirect(nxt)
        _record_login_failure(client_key)
        flash("Akun atau password salah.", "error")
        return render_template("admin/login.html", accounts=accounts, selected_account_key=account_key)
    return render_template("admin/login.html", accounts=accounts, selected_account_key=accounts[0]["key"] if accounts else "")


@app.route("/admin/logout", methods=["POST"])
@login_required
def admin_logout():
    session.clear()
    return redirect(url_for("index"))


@app.route("/admin")
@login_required
def admin_dashboard():
    teams, matches, config = data_context()
    if session.get("admin_role") == "sport":
        admin_sport = session.get("admin_sport")
        matches = [m for m in matches if m.get("sport_key", "table-tennis") == admin_sport]
    enriched = [enrich_match(m, teams) for m in matches]
    enriched.sort(key=utils.match_sort_key)
    counts = {
        "scheduled": sum(1 for m in matches if m["status"] == "scheduled"),
        "live": sum(1 for m in matches if m["status"] == "live"),
        "completed": sum(1 for m in matches if m["status"] == "completed"),
        "postponed": sum(1 for m in matches if m["status"] == "postponed"),
    }
    return render_template("admin/dashboard.html", matches=enriched, counts=counts)


@app.route("/admin/categories")
@login_required
def admin_categories():
    teams = utils.load_teams()
    config = utils.load_config()
    sports = utils.list_sports()
    categories = config.get("categories", [])
    return render_template("admin/categories.html", teams=teams, config=config, sports=sports, categories=categories)


@app.route("/admin/generator")
@login_required
def admin_schedule_generator():
    teams = utils.load_teams()
    config = utils.load_config()
    categories = config.get("categories", [])
    return render_template("admin/schedule_generator.html", teams=teams, config=config, categories=categories)


@app.route("/admin/settings")
@login_required
def admin_settings():
    config = utils.load_config()
    return render_template("admin/settings.html", config=config)


@app.route("/admin/utilities")
@login_required
def admin_utilities():
    config = utils.load_config()
    storage_backend = "legacy"
    return render_template("admin/utilities.html", config=config, storage_backend=storage_backend)



@app.route("/admin/scorekeeper")
@login_required
def scorekeeper_index():
    teams, matches, config = data_context()
    admin_sport_scope = session.get("admin_sport") if session.get("admin_role") == "sport" else None
    if admin_sport_scope:
        matches = [m for m in matches if m.get("sport_key", "table-tennis") == admin_sport_scope]
    sport_filter = admin_sport_scope or request.args.get("sport", "").strip()
    date_filter = request.args.get("date", "").strip()
    court_filter = request.args.get("court", "").strip()
    status_filter = request.args.get("status", "").strip()
    filtered = list(matches)
    if sport_filter:
        filtered = [
            match for match in filtered
            if match.get("sport_key", "table-tennis") == sport_filter
        ]
    if date_filter:
        filtered = [match for match in filtered if match.get("date") == date_filter]
    if court_filter:
        filtered = [match for match in filtered if match.get("court") == court_filter]
    if status_filter:
        filtered = [match for match in filtered if match.get("status") == status_filter]

    status_order = {
        "live": 0, "check_in": 1, "scheduled": 2, "postponed": 3,
        "suspended": 4, "completed": 5, "cancelled": 6,
    }
    filtered.sort(
        key=lambda match: (
            status_order.get(match.get("status"), 9),
            utils.match_sort_key(match),
        )
    )
    cards = []
    for match in filtered:
        document = scorekeeper_document(match, teams)
        cards.append(document)

    counts = Counter(match.get("status") for match in matches)
    sports = [sport for sport in utils.list_sports() if sport.get("enabled")]
    if admin_sport_scope:
        sports = [sport for sport in sports if sport["key"] == admin_sport_scope]
    dates = sorted({match.get("date") for match in matches if match.get("date")})
    courts = sorted({match.get("court") for match in matches if match.get("court")})
    return render_template(
        "admin/scorekeeper_index.html",
        cards=cards,
        counts=counts,
        filter_sports=sports,
        dates=dates,
        courts=courts,
        sport_filter=sport_filter,
        date_filter=date_filter,
        court_filter=court_filter,
        status_filter=status_filter,
    )


@app.route("/admin/scorekeeper/<match_id>")
@login_required
def scorekeeper_match(match_id):
    teams, matches, config = data_context()
    match = utils.get_match(matches, match_id)
    if not match:
        abort(404)
    _assert_sport_access(match.get("sport_key", "table-tennis"))
    document = scorekeeper_document(match, teams)
    return render_template("admin/scorekeeper_match.html", **document)


@app.route("/admin/scorekeeper/<match_id>/action", methods=["POST"])
@login_required
def scorekeeper_action(match_id):
    action = request.form.get("action", "")
    side = request.form.get("side") or None
    reason = request.form.get("reason", "")

    if session.get("admin_role") == "sport":
        current_match = utils.get_match(utils.load_matches(), match_id)
        if not current_match:
            return jsonify({"ok": False, "error": "Pertandingan tidak ditemukan."}), 404
        if current_match.get("sport_key", "table-tennis") != session.get("admin_sport"):
            return jsonify({"ok": False, "error": "Anda tidak memiliki akses ke pertandingan ini."}), 403

    try:
        expected_version = _expected_match_version()
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc), "reload_required": True}), 400

    def apply_action(current):
        utils.apply_scorekeeper_action(
            current, action, side=side, reason=reason
        )

    try:
        updated = utils.update_match(
            match_id, apply_action, expected_version=expected_version
        )
    except utils.MatchVersionConflictError:
        teams, matches, _ = data_context()
        current = utils.get_match(matches, match_id)
        return jsonify(
            {
                "ok": False,
                "error": "Skor berubah di perangkat lain. Data terbaru sudah dimuat; periksa sebelum melanjutkan.",
                "reload_required": True,
                "data": scorekeeper_document(current, teams) if current else None,
            }
        ), 409
    except utils.MatchNotFoundError:
        return jsonify({"ok": False, "error": "Pertandingan tidak ditemukan."}), 404
    except ValueError as exc:
        teams, matches, _ = data_context()
        current = utils.get_match(matches, match_id)
        return jsonify(
            {
                "ok": False,
                "error": str(exc),
                "data": scorekeeper_document(current, teams) if current else None,
            }
        ), 422

    teams = utils.load_teams()
    return jsonify(
        {
            "ok": True,
            "message": {
                "start": "Pertandingan dimulai dan skor Live aktif.",
                "point": "Skor tersimpan.",
                "undo": "Skor terakhir berhasil di-undo.",
                "finish": "Hasil pertandingan dikonfirmasi selesai.",
                "open_correction": "Hasil dibuka untuk koreksi.",
            }.get(action, "Perubahan tersimpan."),
            "data": scorekeeper_document(updated, teams),
        }
    )


@app.route("/admin/reset", methods=["POST"])
@login_required
def admin_reset():
    if utils.is_normalized_backend():
        flash(
            "Reset koleksi legacy dinonaktifkan pada backend normalized. Gunakan import terkontrol dengan backup database.",
            "error",
        )
        return redirect(url_for("admin_dashboard"))
    import generate_data
    utils.backup_data_files("teams.json", "matches.json")
    utils.save_json("teams.json", generate_data.TEAMS)
    utils.save_matches(generate_data.build_matches())
    flash("Semua data pertandingan telah direset ke kondisi awal. Data sebelumnya tersimpan otomatis sebagai cadangan.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/shuffle-putra", methods=["POST"])
@login_required
def admin_shuffle_putra():
    if utils.is_normalized_backend():
        flash(
            "Pengundian ulang legacy dinonaktifkan pada backend normalized.", "error"
        )
        return redirect(url_for("admin_dashboard"))
    import generate_data
    utils.backup_data_files("teams.json", "matches.json")
    teams, matches, config = data_context()
    matches, teams = generate_data.shuffle_putra_groups(matches, teams)
    utils.save_json("teams.json", teams)
    utils.save_matches(matches)
    flash("Grup A dan Grup B Ganda Putra berhasil diundi ulang — lawan tanding tiap tim berubah. Skor laga Ganda Putra yang sudah diinput ikut ter-reset karena pasangannya berubah.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/shuffle-campuran", methods=["POST"])
@login_required
def admin_shuffle_campuran():
    if utils.is_normalized_backend():
        flash(
            "Pengundian ulang legacy dinonaktifkan pada backend normalized.", "error"
        )
        return redirect(url_for("admin_dashboard"))
    import generate_data
    utils.backup_data_files("matches.json")
    _, matches, _ = data_context()
    matches = generate_data.shuffle_campuran_group(matches)
    utils.save_matches(matches)
    flash("Jadwal Ganda Campuran berhasil diundi ulang — urutan lawan tiap tim di tiap tanggal berubah. Skor laga Ganda Campuran yang sudah diinput ikut ter-reset karena pasangannya berubah.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/participants/shuffle-groups/<category_key>", methods=["POST"])
@login_required
def admin_shuffle_groups(category_key):
    if utils.is_normalized_backend():
        flash("Pengundian ulang legacy dinonaktifkan pada backend normalized.", "error")
        return redirect(url_for("admin_schedule_generator"))
    utils.backup_data_files("teams.json")
    try:
        assignment = utils.auto_split_groups(category_key)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("admin_schedule_generator"))
    summary = ", ".join(f"Grup {g}: {len(codes)} tim" for g, codes in assignment.items())
    flash(f"Tim di kategori '{category_key}' berhasil diacak & dibagi ke grup. {summary}", "success")
    return redirect(url_for("admin_schedule_generator"))


@app.route("/admin/announcement", methods=["POST"])
@login_required
def admin_announcement():
    title = request.form.get("announcement_title", "").strip()
    body = request.form.get("announcement_text", "").strip()
    utils.update_announcement(title, body)
    flash("Pengumuman berhasil diperbarui.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/live-stream", methods=["POST"])
@login_required
def admin_live_stream():
    embed_url = request.form.get("youtube_embed_url", "").strip()
    title = request.form.get("youtube_embed_title", "").strip()
    live_chat_enabled = request.form.get("live_chat_enabled") == "on"
    utils.update_live_streaming_config(embed_url, title, live_chat_enabled)
    flash("Konfigurasi Live Streaming YouTube berhasil disimpan.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/participants", methods=["GET"])
@login_required
def admin_participants():
    teams = utils.load_teams()
    config = utils.load_config()
    sports = utils.list_sports()
    categories = config.get("categories", [])
    sites = config.get("sites", [])
    site_lookup = {s["code"]: s["name"] for s in sites}
    return render_template(
        "admin/participants.html", teams=teams, config=config, sports=sports,
        categories=categories, sites=sites, site_lookup=site_lookup,
    )


@app.route("/admin/participants/save", methods=["POST"])
@login_required
def admin_save_participant():
    site_code = request.form.get("site_code", "").strip().upper()
    category = request.form.get("category", "").strip()
    group = request.form.get("group", "").strip()

    if not site_code or not category:
        flash("Site dan kategori divisi wajib dipilih.", "error")
        return redirect(url_for("admin_participants"))

    config = utils.load_config()
    cat_def = next((c for c in config.get("categories", []) if c.get("key") == category), None)
    if not cat_def:
        flash("Kategori divisi tidak ditemukan.", "error")
        return redirect(url_for("admin_participants"))
    if not any(s.get("code") == site_code for s in config.get("sites", [])):
        flash("Site tidak ditemukan. Tambahkan site terlebih dahulu di menu Sites.", "error")
        return redirect(url_for("admin_participants"))

    try:
        code = utils.generate_team_code(site_code, category, config)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("admin_participants"))

    teams = utils.load_teams()
    existing = teams.get(code, {})
    participants = utils.load_participants()

    # ── Kategori tipe Roster (Cadangan): satu record per site berisi daftar pemain cadangan ──
    if cat_def.get("entrant_type") == "roster":
        names = request.form.getlist("reserve_name[]")
        id_numbers = request.form.getlist("reserve_id_number[]")
        roles = request.form.getlist("reserve_role[]")
        reserve_pids = request.form.getlist("reserve_participant_id[]")
        members = []
        for i, (name, id_number, role) in enumerate(zip(names, id_numbers, roles)):
            name = name.strip()
            if not name:
                continue
            existing_pid = reserve_pids[i] if i < len(reserve_pids) else ""
            pid = utils.upsert_participant(
                participants, site_code, name, id_number.strip(),
                existing_id=existing_pid or None,
            )
            members.append({"participant_id": pid, "role": role.strip() or "putra"})
        if not members:
            flash("Minimal 1 pemain cadangan dengan nama harus diisi.", "error")
            return redirect(url_for("admin_participants"))
        utils.save_participants(participants)
        teams[code] = {
            "site_code": site_code,
            "category": category,
            "entrant_type": "roster",
            "members": members,
        }
        utils.save_teams(teams)
        flash(f"Roster cadangan '{code}' ({len(members)} pemain) berhasil disimpan.", "success")
        return redirect(url_for("admin_participants"))

    player1 = request.form.get("player1", "").strip()
    player2 = request.form.get("player2", "").strip()
    color = request.form.get("color", "#2a78d6").strip()
    text = request.form.get("text", "#ffffff").strip()
    id_number1 = request.form.get("id_number1", "").strip()
    email1 = request.form.get("email1", "").strip()
    id_number2 = request.form.get("id_number2", "").strip()
    email2 = request.form.get("email2", "").strip()

    if not player1:
        flash("Minimal Nama Pemain 1 wajib diisi.", "error")
        return redirect(url_for("admin_participants"))

    valid_groups = cat_def.get("groups") or ["A"]
    if group not in valid_groups:
        group = valid_groups[0]

    # ── Photo upload helper (compressed, saved to local disk) ──
    def _upload_photo(file_field_name, slug):
        photo_file = request.files.get(file_field_name)
        if not photo_file or photo_file.filename == "":
            return existing.get(slug)          # keep old photo if nothing uploaded
        allowed = {"jpg", "jpeg", "png", "webp", "gif"}
        ext = photo_file.filename.rsplit(".", 1)[-1].lower() if "." in photo_file.filename else ""
        if ext not in allowed:
            flash(f"Format foto tidak didukung: .{ext}. Gunakan JPG, PNG, atau WEBP.", "error")
            return existing.get(slug)
        url, err = compress_and_upload_image(photo_file, f"participants/{code}_{slug}")
        if err:
            flash(f"Gagal menyimpan foto: {err}", "error")
            return existing.get(slug)
        return url

    photo1 = _upload_photo("photo1", "photo1")
    photo2 = _upload_photo("photo2", "photo2")

    pid1 = utils.upsert_participant(
        participants, site_code, player1, id_number1, email1, photo1 or "",
        existing_id=request.form.get("participant_id1") or None,
    )
    members = [{"participant_id": pid1, "role": "pemain1"}]
    if player2:
        pid2 = utils.upsert_participant(
            participants, site_code, player2, id_number2, email2, photo2 or "",
            existing_id=request.form.get("participant_id2") or None,
        )
        members.append({"participant_id": pid2, "role": "pemain2"})

    reserve_name = request.form.get("reserve_name", "").strip()
    reserve_id_number = request.form.get("reserve_id_number", "").strip()
    if reserve_name:
        pid3 = utils.upsert_participant(
            participants, site_code, reserve_name, reserve_id_number,
            existing_id=request.form.get("reserve_participant_id") or None,
        )
        members.append({"participant_id": pid3, "role": "cadangan"})

    utils.save_participants(participants)

    teams[code] = {
        "site_code": site_code,
        "category": category,
        "group": group,
        "color": color,
        "text": text,
        "members": members,
    }
    utils.save_teams(teams)
    flash(f"Data tim/peserta '{code}' ({player1}) berhasil disimpan.", "success")
    return redirect(url_for("admin_participants"))



@app.route("/admin/participants/delete/<code>", methods=["POST"])
@login_required
def admin_delete_participant(code):
    teams = utils.load_teams()
    if code in teams:
        del teams[code]
        utils.save_teams(teams)
        flash(f"Tim '{code}' berhasil dihapus dari daftar peserta.", "success")
    else:
        flash(f"Tim '{code}' tidak ditemukan.", "error")
    return redirect(url_for("admin_participants"))


@app.route("/admin/categories/save", methods=["POST"])
@login_required
def admin_save_category():
    key = request.form.get("cat_key", "").strip().lower().replace(" ", "_")
    label = request.form.get("cat_label", "").strip()
    sport_key = request.form.get("sport_key", "table-tennis").strip()
    entrant_type = request.form.get("entrant_type", "pair").strip()
    if entrant_type not in ("pair", "roster"):
        entrant_type = "pair"
    has_final = request.form.get("has_final") == "on"

    if not key or not label:
        flash("Kode Divisi (Key) dan Nama Divisi wajib diisi.", "error")
        return redirect(url_for("admin_categories"))

    if entrant_type == "roster":
        # Roster/Cadangan tidak bertanding: tidak ada grup, jadwal, atau babak final.
        groups = []
        has_final = False
    else:
        try:
            group_count = int(request.form.get("group_count", "2"))
        except ValueError:
            group_count = 2
        group_count = max(1, min(group_count, 8))
        groups = [chr(65 + i) for i in range(group_count)]

    config = utils.load_config()
    categories = config.setdefault("categories", [])
    for cat in categories:
        if cat.get("key") == key:
            cat.update({"label": label, "sport_key": sport_key, "entrant_type": entrant_type, "groups": groups, "has_final": has_final})
            break
    else:
        categories.append({"key": key, "label": label, "sport_key": sport_key, "entrant_type": entrant_type, "groups": groups, "has_final": has_final})

    enabled_sports = config.setdefault("enabled_sports", ["table-tennis"])
    if sport_key not in enabled_sports:
        enabled_sports.append(sport_key)

    utils.save_config(config)
    flash(f"Divisi '{label}' ({sport_key.upper()}) berhasil disahkan & diaktifkan di sistem.", "success")
    return redirect(url_for("admin_categories"))


@app.route("/admin/sites")
@login_required
def admin_sites():
    config = utils.load_config()
    sites = config.get("sites", [])
    teams = utils.load_teams()
    site_usage = {}
    for team in teams.values():
        sc = team.get("site_code")
        if sc:
            site_usage[sc] = site_usage.get(sc, 0) + 1
    return render_template("admin/sites.html", sites=sites, site_usage=site_usage)


@app.route("/admin/sites/save", methods=["POST"])
@login_required
def admin_save_site():
    code = request.form.get("code", "").strip().upper()
    name = request.form.get("name", "").strip()
    if not code or not name:
        flash("Kode Site dan Nama Site wajib diisi.", "error")
        return redirect(url_for("admin_sites"))

    config = utils.load_config()
    sites = config.setdefault("sites", [])
    for site in sites:
        if site.get("code") == code:
            site["name"] = name
            break
    else:
        sites.append({"code": code, "name": name})
    utils.save_config(config)
    flash(f"Site '{code}' ({name}) berhasil disimpan.", "success")
    return redirect(url_for("admin_sites"))


@app.route("/admin/sites/delete/<code>", methods=["POST"])
@login_required
def admin_delete_site(code):
    config = utils.load_config()
    sites = config.get("sites", [])
    teams = utils.load_teams()
    if any(team.get("site_code") == code for team in teams.values()):
        flash(f"Site '{code}' masih dipakai oleh tim/peserta terdaftar, tidak bisa dihapus.", "error")
        return redirect(url_for("admin_sites"))
    new_sites = [s for s in sites if s.get("code") != code]
    if len(new_sites) == len(sites):
        flash(f"Site '{code}' tidak ditemukan.", "error")
    else:
        config["sites"] = new_sites
        utils.save_config(config)
        flash(f"Site '{code}' berhasil dihapus.", "success")
    return redirect(url_for("admin_sites"))


CAPSA_CATEGORY_KEY = "capsa_susun"
CAPSA_SPORT_KEY = "capsa-susun"
CAPSA_PARTICIPANTS = [
    {"name": "Jannata", "entity": "IMN"},
    {"name": "Hendra Gunawan", "entity": "IMU"},
    {"name": "Pawestri", "entity": "IMU"},
    {"name": "Casal", "entity": "ISB"},
    {"name": "Dody Indra", "entity": "IMU"},
    {"name": "YIN", "entity": "ISB"},
    {"name": "Linda", "entity": "ISB"},
    {"name": "Indah H", "entity": "IMU"},
    {"name": "Mitarani", "entity": "IMU"},
    {"name": "Rizki Fore", "entity": "Indis"},
    {"name": "Steven", "entity": "ISB"},
    {"name": "Daffa", "entity": "ISB"},
    {"name": "Pak Joko", "entity": "Indis"},
    {"name": "Faisal", "entity": "Indis"},
    {"name": "Sades W", "entity": "ILSS Babelan"},
    {"name": "Januar A", "entity": "ILSS Babelan"},
    {"name": "Rendy R Z", "entity": "ILSS Babelan"},
    {"name": "Rona Kartiko", "entity": "ILSS Babelan"},
    {"name": "Tito", "entity": "IMU"},
    {"name": "Dayat", "entity": "EMITS"},
    {"name": "Mughira", "entity": "EMITS"},
    {"name": "Andre", "entity": "EMITS"},
    {"name": "Fajar", "entity": "Kalista"},
    {"name": "Tito", "entity": "IMU"},
    {"name": "Dino", "entity": "INVI"},
    {"name": "Aldi", "entity": "INVI"},
    {"name": "Raihan", "entity": "INVI"},
    {"name": "Galang", "entity": "INVI"},
    {"name": "Alif", "entity": "INVI"},
    {"name": "Kana", "entity": "INVI"},
    {"name": "Henry", "entity": "INVI"},
    {"name": "Billy", "entity": "INVI"},
    {"name": "Pak Adi Shima", "entity": "IMU"},
    {"name": "Gusna A", "entity": "IMN"},
]
CAPSA_TEAM_COLORS = [
    ("#475569", "#ffffff"), ("#7c3aed", "#ffffff"), ("#0e7490", "#ffffff"),
    ("#b45309", "#ffffff"), ("#be123c", "#ffffff"), ("#166534", "#ffffff"),
]


@app.route("/admin/capsa/pasangan")
@login_required
def admin_capsa_pairs():
    teams = utils.load_teams()
    matches = utils.load_matches()
    pairs = {
        code: t for code, t in teams.items()
        if t.get("category") == CAPSA_CATEGORY_KEY
    }
    used_codes = {
        m.get(slot) for m in matches
        if m.get("category") == CAPSA_CATEGORY_KEY
        for slot in ("team_a", "team_b")
        if m.get(slot)
    }
    capsa_matches = sorted(
        (m for m in matches if m.get("category") == CAPSA_CATEGORY_KEY),
        key=lambda m: m.get("id", ""),
    )
    return render_template(
        "admin/capsa_pairs.html", pairs=pairs, used_codes=used_codes,
        capsa_matches=capsa_matches, teams=teams,
    )


@app.route("/admin/capsa/pasangan/save", methods=["POST"])
@login_required
def admin_save_capsa_pair():
    player1 = request.form.get("player1", "").strip()
    entity1 = request.form.get("entity1", "").strip()
    player2 = request.form.get("player2", "").strip()
    entity2 = request.form.get("entity2", "").strip()
    if not player1 or not player2:
        flash("Nama kedua pemain wajib diisi.", "error")
        return redirect(url_for("admin_capsa_pairs"))

    label1 = f"{player1} ({entity1})" if entity1 else player1
    label2 = f"{player2} ({entity2})" if entity2 else player2

    teams = utils.load_teams()
    existing_nums = [
        int(code.split("-")[1]) for code in teams
        if code.startswith("CAPSA-") and code.split("-")[1].isdigit()
    ]
    next_num = max(existing_nums, default=0) + 1
    code = f"CAPSA-{next_num:02d}"
    color, text = CAPSA_TEAM_COLORS[(next_num - 1) % len(CAPSA_TEAM_COLORS)]

    teams[code] = {
        "category": CAPSA_CATEGORY_KEY,
        "sport_key": CAPSA_SPORT_KEY,
        "group": "EXHIBITION",
        "color": color,
        "text": text,
        "player1": label1,
        "player2": label2,
    }
    utils.save_teams(teams)
    flash(f"Pasangan '{code}' ({label1} / {label2}) berhasil didaftarkan.", "success")
    return redirect(url_for("admin_capsa_pairs"))


@app.route("/admin/capsa/pasangan/delete/<code>", methods=["POST"])
@login_required
def admin_delete_capsa_pair(code):
    teams = utils.load_teams()
    matches = utils.load_matches()
    if any(
        m.get("category") == CAPSA_CATEGORY_KEY and code in (m.get("team_a"), m.get("team_b"))
        for m in matches
    ):
        flash(f"Pasangan '{code}' sudah ditempatkan di bagan, tidak bisa dihapus.", "error")
        return redirect(url_for("admin_capsa_pairs"))
    if code in teams:
        del teams[code]
        utils.save_teams(teams)
        flash(f"Pasangan '{code}' berhasil dihapus.", "success")
    else:
        flash(f"Pasangan '{code}' tidak ditemukan.", "error")
    return redirect(url_for("admin_capsa_pairs"))


@app.route("/admin/generator/group-to-knockout", methods=["POST"])
@login_required
def admin_generate_group_to_knockout():
    category_key = request.form.get("category_key", "").strip()
    start_date = request.form.get("start_date", "").strip()
    court = request.form.get("court", "Meja 1").strip()
    knockout_format = request.form.get("knockout_format", "final_only").strip()
    try:
        utils.backup_data_files("matches.json")
        g_count, k_count = utils.generate_group_to_knockout_schedule(
            category_key, start_date=start_date or None, court=court, knockout_format=knockout_format
        )
        flash(f"Berhasil meng-generate {g_count} laga babak grup dan {k_count} laga babak knockout untuk divisi '{category_key}' dengan format '{knockout_format}'.", "success")
    except Exception as exc:
        flash(f"Gagal generate jadwal: {str(exc)}", "error")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/generator/sync-knockout/<category_key>", methods=["POST"])
@login_required
def admin_sync_knockout(category_key):
    try:
        updated = utils.auto_seed_knockout(category_key, force=True)
        if updated:
            flash(f"Berhasil menyinkronkan & mengesahkan tim lolos babak knockout untuk divisi '{category_key}'!", "success")
        else:
            flash(f"Data tim lolos knockout untuk divisi '{category_key}' sudah sinkron (atau babak grup belum berstatus completed).", "info")
    except Exception as exc:
        flash(f"Gagal menyinkronkan knockout: {str(exc)}", "error")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/rules", methods=["GET"])
@login_required
def admin_rules():
    rules_by_sport = utils.get_sport_rules()
    sports = utils.list_sports()
    return render_template("admin/rules.html", rules_by_sport=rules_by_sport, sports=sports)


@app.route("/admin/rules/save", methods=["POST"])
@login_required
def admin_save_rules():
    rules_by_sport = utils.get_sport_rules()
    for sport_key in rules_by_sport.keys():
        best_of = request.form.get(f"best_of_{sport_key}", type=int)
        if best_of and best_of % 2 == 1:
            rules_by_sport[sport_key]["best_of"] = best_of
        points = request.form.get(f"points_to_win_{sport_key}", type=int)
        if points is not None:
            rules_by_sport[sport_key]["points_to_win"] = points
        games = request.form.get(f"games_to_win_set_{sport_key}", type=int)
        if games is not None:
            rules_by_sport[sport_key]["games_to_win_set"] = games
        win_by = request.form.get(f"win_by_{sport_key}", type=int)
        if win_by is not None:
            rules_by_sport[sport_key]["win_by"] = win_by
        point_cap = request.form.get(f"point_cap_{sport_key}", type=int)
        if point_cap is not None:
            rules_by_sport[sport_key]["point_cap"] = point_cap
        method = request.form.get(f"game_scoring_method_{sport_key}")
        if method is not None and method != "":
            rules_by_sport[sport_key]["game_scoring_method"] = method
        notes = request.form.get(f"notes_{sport_key}", "").strip()
        rules_by_sport[sport_key]["notes"] = notes

    utils.save_sport_rules(rules_by_sport)
    flash("Aturan dan format pertandingan untuk setiap cabang olahraga berhasil diperbarui.", "success")
    return redirect(url_for("admin_rules"))


@app.route("/admin/pertandingan/<match_id>", methods=["GET", "POST"])
@login_required
def admin_edit_match(match_id):
    teams, matches, config = data_context()
    m = utils.get_match(matches, match_id)
    if not m:
        abort(404)
    _assert_sport_access(m.get("sport_key", "table-tennis"))

    if request.method == "POST":
        action = request.form.get("action")
        try:
            expected_version = _expected_match_version()
        except ValueError as exc:
            flash(str(exc), "error")
            expected_version = None
            action = None

        if action == "save_score":
            best_of = utils.sets_needed_to_win(m) * 2 - 1
            segment_label = utils.scorekeeper_terms(m)["segment_label"]
            sets, parse_errors = _parse_score_form(best_of, segment_label)
            validation = utils.validate_match_segments(m, sets)
            requested_status = request.form.get("status", m["status"])
            errors = parse_errors + list(validation.errors)
            if requested_status not in ("live", "scheduled", "postponed"):
                errors.append("Status pertandingan tidak valid.")
            correction_reason = request.form.get("correction_reason", "").strip()
            if m.get("status") == "completed" and sets != (m.get("sets") or []) and not correction_reason:
                errors.append("Alasan koreksi wajib diisi saat mengubah hasil pertandingan selesai.")

            if errors:
                for message in errors:
                    flash(message, "error")
            else:
                notes = request.form.get("notes", "").strip()

                def update_score(current):
                    if not current.get("team_a") or not current.get("team_b"):
                        raise ValueError("Kedua tim harus ditentukan sebelum skor disimpan.")
                    before_sets = current.get("sets") or []
                    if current.get("status") == "completed" and before_sets != sets:
                        current.setdefault("score_corrections", []).append({
                            "before": before_sets,
                            "after": sets,
                            "reason": correction_reason,
                            "actor": "admin",
                            "at": utils.now_wib().isoformat(timespec="seconds"),
                        })
                    current["sets"] = sets
                    current["notes"] = notes
                    current["walkover"] = False
                    if validation.winner_side == "a":
                        current["winner"] = current["team_a"]
                    elif validation.winner_side == "b":
                        current["winner"] = current["team_b"]
                    else:
                        current["winner"] = None
                    current["status"] = "completed" if current["winner"] else requested_status

                _update_match_or_flash(
                    match_id, expected_version, update_score,
                    "Skor pertandingan tersimpan.",
                )

        elif action == "save_capsa_score":
            segments, breakdown, parse_errors = _parse_capsa_score_form()
            validation = utils.validate_match_segments(m, segments)
            requested_status = request.form.get("status", m["status"])
            errors = parse_errors + list(validation.errors)
            if requested_status not in ("live", "scheduled", "postponed"):
                errors.append("Status pertandingan tidak valid.")
            correction_reason = request.form.get("correction_reason", "").strip()
            if m.get("status") == "completed" and segments != (m.get("sets") or []) and not correction_reason:
                errors.append("Alasan koreksi wajib diisi saat mengubah hasil pertandingan selesai.")

            if errors:
                for message in errors:
                    flash(message, "error")
            else:
                notes = request.form.get("notes", "").strip()

                def update_capsa_score(current):
                    if not current.get("team_a") or not current.get("team_b"):
                        raise ValueError("Kedua pasangan harus ditentukan sebelum skor disimpan.")
                    before_sets = current.get("sets") or []
                    if current.get("status") == "completed" and before_sets != segments:
                        current.setdefault("score_corrections", []).append({
                            "before": before_sets,
                            "after": segments,
                            "reason": correction_reason,
                            "actor": "admin",
                            "at": utils.now_wib().isoformat(timespec="seconds"),
                        })
                    current["sets"] = segments
                    current["capsa_rounds"] = breakdown
                    current["notes"] = notes
                    current["walkover"] = False
                    if validation.winner_side == "a":
                        current["winner"] = current["team_a"]
                    elif validation.winner_side == "b":
                        current["winner"] = current["team_b"]
                    else:
                        current["winner"] = None
                    current["status"] = "completed" if current["winner"] else requested_status

                _update_match_or_flash(
                    match_id, expected_version, update_capsa_score,
                    "Skor Capsa Susun tersimpan.",
                )

        elif action == "walkover":
            wo_winner = request.form.get("wo_winner")
            wo_reason = request.form.get("wo_reason", "").strip()
            if wo_winner in (m["team_a"], m["team_b"]):
                def set_walkover(current):
                    if wo_winner not in (current.get("team_a"), current.get("team_b")):
                        raise ValueError("Pilih pemenang WO terlebih dahulu.")
                    current["winner"] = wo_winner
                    current["status"] = "completed"
                    current["walkover"] = True
                    current["sets"] = []
                    current["notes"] = wo_reason or "Menang WO (walkover) — lawan tidak hadir/mengundurkan diri."

                _update_match_or_flash(
                    match_id, expected_version, set_walkover,
                    "Pertandingan ditandai selesai via WO.",
                )
            else:
                flash("Pilih pemenang WO terlebih dahulu.", "error")

        elif action == "set_status":
            requested_status = request.form.get("status")

            def set_status(current):
                if requested_status not in ("scheduled", "live", "postponed"):
                    raise ValueError("Status pertandingan tidak valid.")
                if requested_status == "live" and (
                    not current.get("team_a")
                    or not current.get("team_b")
                    or not current.get("court")
                ):
                    raise ValueError("Pertandingan Live memerlukan dua tim dan meja/lapangan.")
                current["status"] = requested_status

            _update_match_or_flash(
                match_id, expected_version, set_status,
                "Status pertandingan diperbarui.",
            )

        elif action == "reschedule":
            new_date = request.form.get("new_date", "").strip()
            new_time = request.form.get("new_time", "").strip()
            new_court = request.form.get("new_court", "").strip()
            reason = request.form.get("reason", "").strip()
            try:
                datetime.strptime(new_date, "%Y-%m-%d")
                datetime.strptime(new_time, "%H:%M")
            except ValueError:
                flash("Tanggal atau jam baru tidak valid.", "error")
            else:
                if not new_court:
                    flash("Meja/lapangan baru wajib diisi.", "error")
                else:
                    def reschedule_match(current):
                        current.setdefault("reschedule_history", []).append({
                            "from_date": current["date"],
                            "from_time": current["time"],
                            "from_court": current["court"],
                            "to_date": new_date,
                            "to_time": new_time,
                            "to_court": new_court,
                            "reason": reason,
                            "at": utils.now_wib().isoformat(timespec="seconds"),
                        })
                        current["date"], current["time"], current["court"] = (
                            new_date, new_time, new_court
                        )

                    _update_match_or_flash(
                        match_id, expected_version, reschedule_match,
                        "Jadwal pertandingan berhasil diubah (reschedule).",
                    )

        elif action == "set_teams":
            team_a = request.form.get("team_a") or None
            team_b = request.form.get("team_b") or None

            def set_final_teams(current):
                if current.get("stage_type") in {None, "group", "round_robin"} \
                        and current.get("group") != "FINAL":
                    raise ValueError(
                        "Peserta hanya dapat diatur manual untuk stage eliminasi."
                    )
                if team_a and team_a == team_b:
                    raise ValueError("Tim A dan Tim B harus berbeda.")
                eligible = {
                    code for code, team in teams.items()
                    if team.get("category") == current.get("category")
                    and team.get("sport_key", "table-tennis")
                    == current.get("sport_key", "table-tennis")
                }
                if (team_a and team_a not in eligible) or (
                    team_b and team_b not in eligible
                ):
                    raise ValueError(
                        "Peserta stage eliminasi harus berasal dari divisi yang sama."
                    )
                current["team_a"], current["team_b"] = team_a, team_b

            _update_match_or_flash(
                match_id, expected_version, set_final_teams,
                "Tim final diperbarui.",
            )
            
        elif action == "upload_doc":
            file = request.files.get("doc_file") or request.files.get("doc_file_cam")
            if not file or not file.filename:
                flash("Pilih file gambar terlebih dahulu.", "error")
            else:
                file_url, error = compress_and_upload_image(file, match_id)
                if error:
                    flash(f"Gagal mengunggah dokumen: {error}", "error")
                elif file_url:
                    def add_document(current):
                        if len(current.get("docs", [])) >= 3:
                            raise ValueError("Maksimal 3 foto dokumentasi per pertandingan.")
                        current.setdefault("docs", []).append({
                            "url": file_url,
                            "uploaded_at": utils.now_wib().isoformat(timespec="seconds"),
                        })

                    updated = _update_match_or_flash(
                        match_id, expected_version, add_document,
                        "Dokumen berhasil diunggah dan dikompres.",
                    )
                    if updated is None:
                        delete_from_supabase(file_url)
        
        elif action == "delete_doc":
            doc_url = request.form.get("doc_url", "").strip()

            def remove_document(current):
                old_docs = current.get("docs", [])
                new_docs = [doc for doc in old_docs if doc.get("url") != doc_url]
                if not doc_url or len(new_docs) == len(old_docs):
                    raise ValueError("Dokumen tidak ditemukan.")
                current["docs"] = new_docs

            updated = _update_match_or_flash(
                match_id, expected_version, remove_document,
                "Dokumen berhasil dihapus beserta filenya.",
            )
            if updated is not None:
                delete_from_supabase(doc_url)

        elif action == "rotate_doc":
            doc_url = request.form.get("doc_url")
            if doc_url and "docs" in m:
                try:
                    clean_url = doc_url.split("?")[0]
                    local_path = _local_upload_path(clean_url)
                    if not local_path or not os.path.exists(local_path):
                        raise ValueError("Dokumen tidak ditemukan di penyimpanan lokal.")
                    with open(local_path, "rb") as f:
                        body = f.read()
                    img = Image.open(io.BytesIO(body))
                    img.verify()
                    img = Image.open(io.BytesIO(body))
                    img = img.rotate(-90, expand=True)
                    if img.mode != "RGB":
                        img = img.convert("RGB")

                    img.save(local_path, format="JPEG", quality=75, optimize=True)

                    rotated_url = clean_url + f"?v={utils.now_wib().timestamp()}"

                    def rotate_document(current):
                        for document in current.get("docs", []):
                            if document.get("url", "").split("?")[0] == clean_url:
                                document["url"] = rotated_url
                                return
                        raise ValueError("Dokumen tidak ditemukan.")

                    _update_match_or_flash(
                        match_id, expected_version, rotate_document,
                        "Dokumen berhasil di-rotate 90 derajat.",
                    )
                except Exception as e:
                    flash(f"Gagal me-rotate dokumen: {e}", "error")

        elif action is not None:
            flash("Aksi admin tidak dikenal.", "error")

        if request.form.get("modal"):
            return redirect(url_for("match_fragment", match_id=match_id))
        return redirect(url_for("admin_edit_match", match_id=match_id))

    qualification = qualification_context(m, matches, teams, config)

    return render_template(
        "admin/edit_match.html", m=enrich_match(m, teams), raw=m,
        teams=teams, **qualification,
    )


# ---------- license activation routes ----------

@app.route("/license-lockout")
def license_lockout():
    reason = request.args.get("reason", "Lisensi website belum aktif atau masa berlakunya telah habis sesuai ketentuan dari Berlanggan.web.id.")
    return render_template("license_lockout.html", reason=reason), 403


@app.route("/admin/license")
@login_required
def admin_license():
    lic = license_client.validate(force=False)
    return render_template("admin/license.html", license=lic)


@app.route("/admin/license/activate", methods=["POST"])
@login_required
def admin_activate_license():
    key = request.form.get("license_key", "").strip()
    success, message, code = license_client.activate(key)
    if success:
        flash(message, "success")
    else:
        flash(message, "error")
    return redirect(url_for("admin_license"))


@app.route("/admin/license/validate", methods=["POST"])
@login_required
def admin_validate_license():
    lic = license_client.validate(force=True)
    if lic.get("status") in ("active", "grace"):
        flash("Sync & Validasi ke Berlanggan.web.id berhasil! Lisensi aktif.", "success")
    else:
        flash(f"Status lisensi: {lic.get('status').upper()}. {lic.get('last_message')}", "warning")
    return redirect(url_for("admin_license"))


@app.route("/admin/license/deactivate", methods=["POST"])
@login_required
def admin_deactivate_license():
    success, message = license_client.deactivate()
    flash(message, "info")
    return redirect(url_for("admin_license"))


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
